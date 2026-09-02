from pathlib import Path

from openpyxl import Workbook

from gos.extensions import db
from gos.modulos.recursos.importer import import_planilla
from gos.modulos.recursos.models import RecUnidad
from gos.modulos.recursos.services import (
    asignar_destino,
    crear_unidad,
    dar_de_baja,
    detalle_tablero,
    listar_unidades,
    resumen,
)


def _sample_xlsx(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Unidades"
    ws["H2"] = "AFECTACION DE UNIDADES POR SERVICIO Y POR ESTRUCTURA"
    ws["H3"] = "ACON-HWO"
    ws["I3"] = "MANT"
    ws["J3"] = "LIBRE"
    ws["G4"] = "UL"
    ws["H4"] = 1
    ws["I4"] = 1
    ws["J4"] = 0
    ws["G5"] = "TR"
    ws["H5"] = 1
    ws["G6"] = "ST"
    ws["G7"] = "OT"
    ws["C17"] = "INTERNO"
    ws["D17"] = "DOMINIO"
    ws["G17"] = "CO"
    ws["H17"] = "H18"
    ws["I17"] = "GOS"
    ws["J17"] = "GOS"
    ws["C19"] = "UL 01"
    ws["D19"] = "ABC 123"
    ws["G19"] = "OP"
    ws["H19"] = "O"
    ws["C20"] = "TR 02"
    ws["D20"] = "DEF 456"
    ws["I20"] = "O"
    ws["C21"] = "ST 03"
    ws["D21"] = "GHI 789"
    ws["C22"] = "UL 04"
    ws["D22"] = "JKL 000"
    ws["J22"] = "O"
    ws["C23"] = "H18"
    ws["D23"] = "VAN CTR"
    ws["H23"] = "O"
    wb.save(path)
    return path


def test_recursos_paginas(auth_client):
    for url in (
        "/gos/recursos/",
        "/gos/recursos/flota",
        "/gos/recursos/servicios",
        "/gos/recursos/centros",
        "/gos/recursos/importar",
    ):
        r = auth_client.get(url)
        assert r.status_code == 200, url


def test_tablero_lineas_clickeables(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        data = resumen()
        det = detalle_tablero(data)
        ul = det["tipos"]["UL"]
        assert ul["total"] == 2
        acon_ul = next(d for d in ul["destinos"] if "ACON" in (d["nombre"] or ""))
        assert any(u["interno"] == "UL 01" for u in acon_ul["unidades"])
        serv = det["grupos"]["servicio"]
        assert serv["total"] == 2
        assert any("UL 01" in [u["interno"] for u in s["unidades"]] for s in serv["destinos"])
        assert det["grupos"]["none"]["total"] == 1
        hueco_tr = next(h for h in det["huecos"] if h["tipo"] == "TR")
        assert hueco_tr["faltan"] == 1
        assert hueco_tr["destino_id"]

    r = auth_client.get("/gos/recursos/")
    assert r.status_code == 200
    html = r.data.decode("utf-8", "replace")
    assert 'data-rec-kind="hueco"' in html
    assert 'data-rec-kind="tipo"' in html
    assert 'data-rec-kind="grupo"' in html
    assert 'id="rec-detalle"' in html
    assert "modal-dialog-centered" in html
    assert "offcanvas" not in html
    assert "Historial de cambios" in html
    assert "Unidad liviana" in html
    assert "Orden de Trabajo" in html
    assert "UL 01" in html


def test_importar_planilla_y_cambiar_destino(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        result = import_planilla(xlsx, db.session)
        assert result["unidades"] == 5
        assert result["asignadas"] == 4
        assert result["sin_asignar"] == 1
        assert result["destinos"] == 3

        data = resumen()
        assert data["unidades"] == 5
        assert data["sin_asignar"] == 1
        assert data["libres"] == 1
        assert data["por_tipo"]["UL"] == 2
        assert data["por_tipo"]["TR"] == 1
        assert data["por_tipo"]["OT"] == 1
        assert data["por_grupo"]["servicio"] == 2
        assert data["por_grupo"]["estructura"] == 1
        assert any(d["codigo"] == "MANT" for d in data["estructura"])
        assert any(d["codigo"] == "LIBRE" for d in data["estados"])
        assert data["total"]["codigo"] == "TOTAL"
        assert data["total"]["afectadas"] == 2
        assert data["total"]["tipos"][0]["necesarias"] == 1
        assert [c["codigo"] for c in data["parque_columnas"]] == ["MANT", "LIBRE", "TOTAL"]

        h18 = RecUnidad.query.filter_by(codigo="H18").one()
        assert h18.es_centro is True
        acon = next(s for s in data["servicios"] if s.get("equipo") == "H18")
        assert acon["afectadas"] == 1
        assert next(t for t in acon["tipos"] if t["tipo"] == "UL")["afectadas"] == 1
        assert next(t for t in acon["tipos"] if t["tipo"] == "OT")["afectadas"] == 0
        assert any(u["interno"] == "UL 01" for u in acon["flota"])
        assert any(u["interno"] == "ST 03" for u in data["sin_asignar_flota"])

        orden = [u.codigo for u in listar_unidades()]
        assert orden[0] == "ST03"
        assert "UL04" in orden[1:]

    r = auth_client.get("/gos/recursos/servicios")
    assert r.status_code == 200
    assert b"Estructura y estado de parque" in r.data
    assert b"MANT" in r.data
    assert b"LIBRE" in r.data
    assert b"TOTAL" in r.data
    assert b"rec-chip" in r.data
    assert b"UL 01" in r.data
    assert b"Sin asignar" in r.data
    assert b"rec-chip-equipo" in r.data
    assert b'data-kind="centro"' in r.data
    assert b"rec-parque-neg" in r.data

    with app.app_context():
        huecos_tr = [h for h in resumen()["huecos"] if h["tipo"] == "TR"]
        assert huecos_tr
        assert huecos_tr[0]["faltan"] == 1

        from gos.modulos.recursos.models import RecCentro, RecDestino

        h18c = RecCentro.query.filter_by(codigo="H18").one()
        acon = RecDestino.query.filter_by(equipo="H18").one()
        assert h18c.destino_id == acon.id
        assert h18c.activo is True
        ul01 = RecUnidad.query.filter_by(codigo="UL01").one()
        assert ul01.asignacion is not None
        libre = RecDestino.query.filter_by(codigo="LIBRE").one()
        changed = asignar_destino(ul01.id, libre.id)
        assert changed["destino_codigo"] == "LIBRE"

    r = auth_client.get("/gos/recursos/flota")
    assert r.status_code == 200
    assert b"rec-pendiente" in r.data
    assert b"Agregar unidad" in r.data
    assert b"Dar de baja" in r.data

    with app.app_context():
        ul01 = RecUnidad.query.filter_by(codigo="UL01").one()
        r = auth_client.post(
            f"/gos/recursos/api/unidades/{ul01.id}/destino",
            json={"destino_id": None},
        )
    assert r.status_code == 200
    payload = r.get_json()
    assert payload["ok"] is True
    assert payload["unidad"]["destino_id"] is None


def test_importar_via_api(auth_client, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asig.xlsx")
    with xlsx.open("rb") as fh:
        r = auth_client.post(
            "/gos/recursos/api/importar",
            data={"file": (fh, "planilla.xlsx")},
            content_type="multipart/form-data",
        )
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True
    assert data["detalle"]["unidades"] == 5


def test_alta_baja_y_unidad_de_centro(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        acon = RecDestino.query.filter_by(equipo="H18").one()
        van = crear_unidad(
            interno="VAN 01",
            dominio="VAN 111",
            tipo="OT",
            es_centro=True,
            destino_id=acon.id,
        )
        assert van["es_centro"] is True
        assert van["destino_id"] == acon.id

        data = resumen()
        card = next(s for s in data["servicios"] if s.get("equipo") == "H18")
        assert [c["interno"] for c in card["centros"]] == ["VAN 01"]
        assert card["afectadas"] == 1

        r = auth_client.get("/gos/recursos/servicios")
        assert r.status_code == 200
        assert b"VAN 01" in r.data

        baja = dar_de_baja(van["id"])
        assert baja["activo"] is False
        assert RecUnidad.query.filter_by(codigo="VAN01").one().activo is False
        assert resumen()["unidades"] == 5

        r = auth_client.post(
            "/gos/recursos/api/unidades",
            json={"interno": "UL 88", "dominio": "ZZZ 888", "tipo": "UL"},
        )
    assert r.status_code == 200
    payload = r.get_json()
    assert payload["ok"] is True
    assert payload["unidad"]["interno"] == "UL 88"

    with app.app_context():
        ul88 = RecUnidad.query.filter_by(codigo="UL88").one()
        r = auth_client.post(f"/gos/recursos/api/unidades/{ul88.id}/baja")
        assert r.status_code == 200
        assert RecUnidad.query.filter_by(codigo="UL88").one().activo is False

        import_planilla(xlsx, db.session)
        assert RecUnidad.query.filter_by(codigo="VAN01").one().activo is False
        assert RecUnidad.query.filter_by(codigo="UL88").one().activo is False


def test_reasignar_mueve_unidad_en_servicios_y_flota(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        st03 = RecUnidad.query.filter_by(codigo="ST03").one()
        acon = RecDestino.query.filter_by(equipo="H18").one()
        r = auth_client.post(
            f"/gos/recursos/api/unidades/{st03.id}/destino",
            json={"destino_id": acon.id},
        )
        assert r.status_code == 200
        assert r.get_json()["unidad"]["destino_id"] == acon.id

        data = resumen()
        card = next(s for s in data["servicios"] if s.get("equipo") == "H18")
        assert any(u["interno"] == "ST 03" for u in card["flota"])
        assert not any(u["interno"] == "ST 03" for u in data["sin_asignar_flota"])

    flota = auth_client.get("/gos/recursos/flota")
    assert flota.status_code == 200
    html = flota.get_data(as_text=True)
    assert "ST 03" in html
    assert html.split("ST 03", 1)[1].split("</tr>", 1)[0].count("rec-pendiente") == 0


def test_mover_centro_entre_servicios(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        acon = RecDestino.query.filter_by(equipo="H18").one()
        otro = RecDestino(
            codigo="ACON-HWO-EL",
            nombre="ACON-HWO-EL",
            grupo="servicio",
            equipo="H23",
            orden=1,
            activo=True,
        )
        db.session.add(otro)
        db.session.commit()
        van = crear_unidad(
            interno="VAN 01",
            tipo="OT",
            es_centro=True,
            destino_id=acon.id,
        )
        r = auth_client.post(
            f"/gos/recursos/api/destinos/{otro.id}/equipo",
            json={"equipo": "H18", "desde_id": acon.id},
        )
        assert r.status_code == 200
        payload = r.get_json()
        assert payload["ok"] is True
        assert payload["equipo"] == "H18"
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=otro.id).one().equipo == "H18"
        assert RecDestino.query.filter_by(id=acon.id).one().equipo is None
        data = resumen()
        vacante = next(s for s in data["servicios"] if s["id"] == acon.id)
        assert vacante["sin_centro"] is True
        assert data["servicios"][0]["id"] == acon.id
        html = auth_client.get("/gos/recursos/servicios").get_data(as_text=True)
        assert "rec-alerta-centro" in html
        assert RecUnidad.query.filter_by(codigo="VAN01").one().asignacion.destino_id == otro.id
        assert RecUnidad.query.filter_by(codigo="UL01").one().asignacion.destino_id == acon.id


def test_crear_y_editar_servicio_con_requerimientos(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        r = auth_client.post(
            "/gos/recursos/api/destinos",
            json={
                "nombre": "YPF-HWO-NUEVO",
                "equipo": "H99",
                "cupos": {"UL": 2, "TR": 1, "ST": 0, "OT": 0},
            },
        )
        assert r.status_code == 200
        dest = r.get_json()["destino"]
        assert dest["nombre"] == "YPF-HWO-NUEVO"
        assert dest["equipo"] == "H99"
        assert dest["cupos"]["UL"] == 2
        assert dest["cupos"]["TR"] == 1

        r2 = auth_client.post(
            f"/gos/recursos/api/destinos/{dest['id']}",
            json={"cupos": {"UL": 3, "TR": 1, "ST": 1, "OT": 0}},
        )
        assert r2.status_code == 200
        assert r2.get_json()["destino"]["cupos"]["UL"] == 3
        assert r2.get_json()["destino"]["cupos"]["ST"] == 1

        data = resumen()
        card = next(s for s in data["servicios"] if s["id"] == dest["id"])
        assert next(t for t in card["tipos"] if t["tipo"] == "UL")["necesarias"] == 3
        assert card["sin_centro"] is False

        html = auth_client.get("/gos/recursos/servicios").get_data(as_text=True)
        assert "Agregar servicio" in html
        assert "rec-svc-modal" in html
        assert "rec-parque-edit" in html

        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        vivo = RecDestino.query.filter_by(id=dest["id"]).one()
        assert vivo.activo is True
        assert {c.tipo: c.necesarias for c in vivo.cupos}["UL"] == 3


def test_editar_necesarias_columna_parque(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        mant = RecDestino.query.filter_by(codigo="MANT").one()
        r = auth_client.post(
            f"/gos/recursos/api/destinos/{mant.id}",
            json={"cupos": {"UL": 4, "TR": 2, "ST": 0, "OT": 1}},
        )
        assert r.status_code == 200
        cupos = r.get_json()["destino"]["cupos"]
        assert cupos["UL"] == 4
        assert cupos["TR"] == 2
        assert cupos["OT"] == 1
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=mant.id).one().nombre == "MANT"

        r = auth_client.post(
            f"/gos/recursos/api/destinos/{mant.id}",
            json={"nombre": "HACK"},
        )
        assert r.status_code == 400

        data = resumen()
        col = next(c for c in data["parque_columnas"] if c["codigo"] == "MANT")
        assert next(t for t in col["tipos"] if t["tipo"] == "UL")["necesarias"] == 4
        total = next(c for c in data["parque_columnas"] if c["codigo"] == "TOTAL")
        assert next(t for t in total["tipos"] if t["tipo"] == "UL")["necesarias"] == 4

        html = auth_client.get("/gos/recursos/servicios").get_data(as_text=True)
        assert "rec-parque-edit" in html
        assert f'data-destino-id="{mant.id}"' in html
        assert 'data-nombre="TOTAL"' not in html


def test_baja_servicio_y_centro_sin_asignar(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecCentro, RecDestino, RecUnidad

        acon = RecDestino.query.filter_by(equipo="H18").one()
        ul01 = RecUnidad.query.filter_by(codigo="UL01").one()
        r = auth_client.post(
            f"/gos/recursos/api/destinos/{acon.id}",
            json={"equipo": ""},
        )
        assert r.status_code == 200
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=acon.id).one().equipo is None
        h18 = RecCentro.query.filter_by(codigo="H18").one()
        assert h18.destino_id is None
        data = resumen()
        assert any(c["codigo"] == "H18" for c in data["sin_asignar_centros"])
        vacante = next(s for s in data["servicios"] if s["id"] == acon.id)
        assert vacante["sin_centro"] is True
        html = auth_client.get("/gos/recursos/servicios").get_data(as_text=True)
        assert "H18" in html
        assert "Dar de baja" in html

        r = auth_client.post(
            f"/gos/recursos/api/destinos/{acon.id}",
            json={"equipo": "H18"},
        )
        assert r.status_code == 200
        r = auth_client.post(f"/gos/recursos/api/destinos/{acon.id}/baja")
        assert r.status_code == 200
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=acon.id).one().activo is False
        assert RecCentro.query.filter_by(codigo="H18").one().destino_id is None
        assert RecUnidad.query.filter_by(codigo="UL01").one().asignacion is None
        data = resumen()
        assert all(s["id"] != acon.id for s in data["servicios"])
        assert any(c["codigo"] == "H18" for c in data["sin_asignar_centros"])
        assert any(u["interno"] == "UL 01" for u in data["sin_asignar_flota"])
        assert ul01.id


def test_centros_crud_y_pagina(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecCentro, RecDestino

        r = auth_client.get("/gos/recursos/centros")
        assert r.status_code == 200
        html = r.get_data(as_text=True)
        assert "H18" in html
        assert "ACON-HWO" in html
        assert "Agregar centro" in html

        r = auth_client.post(
            "/gos/recursos/api/centros",
            json={"codigo": "H50", "nombre": "Centro 50"},
        )
        assert r.status_code == 200
        centro = r.get_json()["centro"]
        assert centro["codigo"] == "H50"
        assert centro["sin_asignar"] is True
        assert any(c["codigo"] == "H50" for c in resumen()["sin_asignar_centros"])

        acon = RecDestino.query.filter_by(equipo="H18").one()
        r = auth_client.post(
            f"/gos/recursos/api/centros/{centro['id']}",
            json={"destino_id": acon.id},
        )
        assert r.status_code == 200
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=acon.id).one().equipo == "H50"
        assert RecCentro.query.filter_by(codigo="H18").one().destino_id is None
        assert RecCentro.query.filter_by(codigo="H50").one().destino_id == acon.id

        r = auth_client.post(
            f"/gos/recursos/api/centros/{centro['id']}",
            json={"destino_id": None, "nombre": "Base 50"},
        )
        assert r.status_code == 200
        assert r.get_json()["centro"]["nombre"] == "Base 50"
        db.session.expire_all()
        assert RecDestino.query.filter_by(id=acon.id).one().equipo is None

        r = auth_client.post("/gos/recursos/api/centros/mover", json={"equipo": "H18", "destino_id": acon.id})
        assert r.status_code == 200
        db.session.expire_all()
        assert RecCentro.query.filter_by(codigo="H18").one().destino_id == acon.id

        r = auth_client.post("/gos/recursos/api/centros/mover", json={"equipo": "H18", "destino_id": None})
        assert r.status_code == 200
        db.session.expire_all()
        assert RecCentro.query.filter_by(codigo="H18").one().destino_id is None
        assert RecDestino.query.filter_by(id=acon.id).one().equipo is None

        r = auth_client.post(f"/gos/recursos/api/centros/{centro['id']}/baja")
        assert r.status_code == 200
        db.session.expire_all()
        assert RecCentro.query.filter_by(codigo="H50").one().activo is False
        assert RecCentro.query.filter_by(codigo="H50").one().destino_id is None


def test_historial_quien_fecha_y_cambio(auth_client, app, tmp_path):
    xlsx = _sample_xlsx(tmp_path / "asignacion.xlsx")
    with app.app_context():
        import_planilla(xlsx, db.session)
        from gos.modulos.recursos.models import RecDestino

        st03 = RecUnidad.query.filter_by(codigo="ST03").one()
        acon = RecDestino.query.filter_by(equipo="H18").one()
        r = auth_client.post(
            f"/gos/recursos/api/unidades/{st03.id}/destino",
            json={"destino_id": acon.id},
        )
        assert r.status_code == 200

    html = auth_client.get("/gos/recursos/").get_data(as_text=True)
    assert "Historial de cambios" in html
    assert "rec-historial" in html
    assert "Test" in html
    assert "ST 03" in html
    assert "Sin asignar" in html
    assert "ACON-HWO" in html
    assert "Importó planilla" in html


def test_recursos_en_tablas_protegidas():
    from gos.modulos.objetivos.services.import_service import TABLES

    for table in ("rec_destinos", "rec_centros", "rec_unidades", "rec_cupos", "rec_asignaciones", "rec_cambios"):
        assert table in TABLES
