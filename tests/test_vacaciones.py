import pytest

from gos.modulos.vacaciones import storage
from gos.modulos.vacaciones.models import Registro, Vacacion


@pytest.fixture(autouse=True)
def vacaciones_db(app):
    with app.app_context():
        storage.reset_for_tests()
        yield
        storage.reset_for_tests()


def test_vacaciones_health_requires_auth(client):
    r = client.get("/gos/vacaciones/api/health")
    assert r.status_code == 302
    assert "/auth/login" in r.location


def test_vacaciones_health_ok(auth_client):
    r = auth_client.get("/gos/vacaciones/api/health")
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True


def test_vacaciones_shell(auth_client):
    r = auth_client.get("/gos/vacaciones/")
    assert r.status_code == 200
    assert b"vac-frame" in r.data


def test_vacaciones_app(auth_client):
    r = auth_client.get("/gos/vacaciones/app/")
    assert r.status_code == 200
    assert b"Vacaciones Adeudadas" in r.data


def test_vacaciones_deuda_empty(auth_client):
    r = auth_client.get("/gos/vacaciones/api/vacaciones/deuda")
    assert r.status_code == 200
    assert r.get_json() == []


def test_vacaciones_dashboard_lists(auth_client, app):
    with app.app_context():
        from datetime import date

        from gos.extensions import db

        db.session.add(
            Registro(fecha=date(2025, 6, 1), empleado="Ana Test", sector="IT", vacaciones=0)
        )
        db.session.add(
            Vacacion(
                legajo=1,
                empleado="Ana Test",
                sector="IT",
                anio=2025,
                dias_disponibles=14,
                dias_tomados=5,
                dias_pendientes=9,
            )
        )
        db.session.commit()

    r = auth_client.get("/gos/vacaciones/api/dashboard/años")
    assert r.status_code == 200
    assert 2025 in r.get_json()

    r = auth_client.get("/gos/vacaciones/api/dashboard/sectores")
    assert r.status_code == 200
    assert "IT" in r.get_json()

    r = auth_client.get("/gos/vacaciones/api/dashboard/empleados")
    assert r.status_code == 200
    assert "Ana Test" in r.get_json()


def test_vacaciones_import_excel(auth_client):
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOTAL"
    ws.append([
        "fecha", "empleado", "sector", "servicio", "centro", "situacion",
        "total_horas", "hs_viaje", "hs50", "hs_noc", "hs_noc50", "hs100",
        "viandas", "v_desayuno", "d_normales", "ausente", "fr_trabajados",
        "feriados", "enfermedad", "traslado", "vacaciones", "licencia",
        "suspension", "accidente", "francos_comp",
    ])
    ws.append([
        "2025-06-01", "Pedro Test", "RRHH", "", "", "", 8, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0,
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/gos/vacaciones/api/importar/excel",
        data={"file": (buf, "test.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True
    assert data["detalle"]["registros"] == 1


def test_vacaciones_import_planilla_archivo_actualizado(auth_client, app):
    """Importa el formato real: una hoja tipo planilla (sin TOTAL)."""
    import io
    from datetime import date, datetime

    import openpyxl
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append([
        "N° Legajo", "VACACIONES ", None, None, "Antigüedad ",
        "VACACIONES  2023", None, None, "Antigüedad",
        "VACACIONES  2024", None, None, "Antigüedad",
        "VACACIONES 2025", None, None,
    ])
    ws.append([
        None, "Empleado", "Fecha ingreso", "Sector", None,
        "Días Disponibles", "Días tomados", "Dias pendientes", None,
        "Días disponibles", "Días tomados", "Dias pendientes",
        datetime(2025, 12, 31), "Días Disponibles", "Días tomados", "Dias pendientes",
    ])
    ws.append([
        None, None, None, None, datetime(2023, 12, 31),
        None, None, None, datetime(2024, 12, 31),
        None, None, None, None, None, None, None,
    ])
    ws.append([
        43, "Alias, Fernando Javier", datetime(2007, 1, 1), "Operaciones",
        16, 28, 28, 0, 17, 28, 11, 17, 18, 28, None, 28,
        "Baja 2026", "Nota R test",
    ])
    ws.append([
        79, "Araneda, Ana Romina", datetime(2008, 1, 7), "Operaciones",
        15, 28, 28, 0, 16, 28, 28, 0, 16, 28, 3, 25,
        None, None,
    ])
    # Fórmula de suma en "Días tomados" (antes se leía como 0).
    ws["G4"] = "=SUM(7,21)"
    ws["O5"] = "=7+3+5"
    # Comentarios (triángulo rojo) en Días tomados 2023 (G) y 2025 (O)
    ws["G4"].comment = Comment("Tester:\n(7) 01/01 al 07/01/2023", "Tester")
    ws["O5"].comment = Comment("Tester:\n(3) 01-05-26 al 03-05-26", "Tester")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/gos/vacaciones/api/importar/excel",
        data={"file": (buf, "Archivo de Vacaciones-Actualizado.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True
    assert data["detalle"]["vacaciones"] == 6  # 2 empleados × 3 años
    assert data["detalle"]["registros"] == 0

    with app.app_context():
        from gos.extensions import db

        rows = db.session.query(Vacacion).order_by(Vacacion.empleado, Vacacion.anio).all()
        assert len(rows) == 6
        ana_2025 = next(r for r in rows if r.empleado.startswith("Araneda") and r.anio == 2025)
        assert ana_2025.dias_disponibles == 28
        assert ana_2025.dias_tomados == 15  # =7+3+5
        assert ana_2025.dias_pendientes == 25
        assert ana_2025.fecha_ingreso == date(2008, 1, 7)
        assert ana_2025.comentario == "(3) 01-05-26 al 03-05-26"

        alias_2023 = next(r for r in rows if r.empleado.startswith("Alias") and r.anio == 2023)
        assert alias_2023.dias_tomados == 28  # =SUM(7,21)
        assert alias_2023.comentario == "(7) 01/01 al 07/01/2023"
        assert alias_2023.nota_q == "Baja 2026"
        assert alias_2023.nota_r == "Nota R test"

    r = auth_client.get("/gos/vacaciones/api/vacaciones/deuda")
    assert r.status_code == 200
    deuda = r.get_json()
    assert len(deuda) == 6
    assert any(d["dias_pendientes"] == 17 for d in deuda)
    alias_api = next(d for d in deuda if d["empleado"].startswith("Alias") and d["anio"] == 2023)
    assert alias_api["fecha_ingreso"] == "2007-01-01"
    assert alias_api["comentario"] == "(7) 01/01 al 07/01/2023"
    assert alias_api["nota_q"] == "Baja 2026"
    assert alias_api["nota_r"] == "Nota R test"

    r = auth_client.get("/gos/vacaciones/api/dashboard/sectores")
    assert "Operaciones" in r.get_json()

    r = auth_client.get("/gos/vacaciones/api/dashboard/años")
    assert r.get_json() == [2023, 2024, 2025]


def test_tot_hs_shell_and_app(auth_client):
    r = auth_client.get("/gos/vacaciones/tot-hs")
    assert r.status_code == 200
    assert b"vac-frame" in r.data

    r = auth_client.get("/gos/vacaciones/app/?view=tot_hs")
    assert r.status_code == 200
    assert b"Tot Hs." in r.data
    assert b"view-tot-hs" in r.data
    assert b"Cargar Excel" not in r.data  # la carga vive en Importar datos
    assert b"Importar datos" in r.data


def test_importar_centraliza_cargas(auth_client):
    r = auth_client.get("/gos/vacaciones/app/?view=importar")
    assert r.status_code == 200
    assert b"Importar datos" in r.data
    assert b"Vacaciones adeudadas" in r.data
    assert b"Tot Hs." in r.data
    assert b"id=\"dropzone\"" in r.data
    assert b"id=\"ths-dropzone\"" in r.data


def _xlsx_tot_hs_period(title, rows, sheet_title="Total"):
    """Excel tipo archivo (14): título con rango + Nombre/Servicio/Centro/horas."""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append([title])
    ws.append([
        "Nombre", "Servicio", "Centro", "Cliente", "Tipo Servicio",
        "Total Horas", "Hs.Viaje", "Hs.50", "Hs.Noct", "Hs.Noct50", "Hs.100",
        "Viandas", "Desayunos", "D.Normales", "Ausente", "Fr.Trabajados",
        "Feriado", "Enfermo", "Traslado", "Vacaciones", "Licencia",
        "Accidente", "Fr.Compens", "TotalHs+HsViaje",
    ])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_tot_hs_import_period_merge_and_overwrite(auth_client, app):
    """Períodos nuevos se suman; el mismo rango se reemplaza completo."""
    from datetime import date

    from gos.modulos.vacaciones.models import TotHs

    buf1 = _xlsx_tot_hs_period(
        "21/12/2025 al 20/07/2026",
        [
            ["ALIAS ALEJANDRO DARIO", "PAMPA-PTF-RDA", "PRAF-04", "PAMPA ENERGIA", "PLANTAS",
             100, 10, 5, 2, 1, 3, 4, 1, 20, 0, 0, 1, 0, 2, 0, 0, 0, 0, 110],
            ["ALIAS FERNANDO JAVIER", "GOS-OPE-PCL", "BASE NQN", "GOS", "PCL",
             80, 0, 8, 0, 0, 4, 10, 0, 15, 0, 1, 0, 0, 0, 2, 0, 0, 0, 80],
        ],
    )
    r = auth_client.post(
        "/gos/vacaciones/api/importar/total",
        data={"file": (buf1, "archivo (14).xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200, r.get_json()
    data = r.get_json()
    assert data["ok"] is True
    assert data["detalle"]["registros"] == 2
    assert data["detalle"]["periodo_reemplazado"] is False
    assert data["detalle"]["fecha_min"] == "2025-12-21"
    assert data["detalle"]["fecha_max"] == "2026-07-20"

    # Mismo período: pisa (reemplaza las 2 filas por 1)
    buf2 = _xlsx_tot_hs_period(
        "21/12/2025 al 20/07/2026",
        [
            ["ALIAS ALEJANDRO DARIO", "PAMPA-PTF-RDA", "PRAF-04", "PAMPA ENERGIA", "PLANTAS",
             200, 20, 10, 0, 0, 0, 0, 0, 30, 0, 0, 0, 0, 0, 0, 0, 0, 0, 220],
        ],
        sheet_title="Hoja1",
    )
    r = auth_client.post(
        "/gos/vacaciones/api/importar/total",
        data={"file": (buf2, "archivo-upd.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["detalle"]["periodo_reemplazado"] is True
    assert data["detalle"]["registros"] == 1

    with app.app_context():
        from gos.extensions import db

        rows = db.session.query(TotHs).all()
        assert len(rows) == 1
        assert rows[0].empleado == "ALIAS ALEJANDRO DARIO"
        assert rows[0].total_horas == 200
        assert rows[0].periodo_desde == date(2025, 12, 21)

    # Período distinto: se agrega sin borrar el anterior
    buf3 = _xlsx_tot_hs_period(
        "01/01/2025 al 20/12/2025",
        [
            ["PEZO SERGIO HERALDO", "SHELL-HWO-02", "H21", "SHELL CAPSA", "HOT WATER OIL",
             50, 5, 2, 0, 0, 1, 3, 0, 10, 0, 0, 0, 0, 0, 0, 0, 0, 0, 55],
        ],
    )
    r = auth_client.post(
        "/gos/vacaciones/api/importar/total",
        data={"file": (buf3, "archivo-prev.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    assert r.get_json()["detalle"]["periodo_reemplazado"] is False

    with app.app_context():
        from gos.extensions import db

        assert db.session.query(TotHs).count() == 2

    r = auth_client.get("/gos/vacaciones/api/tot-hs/meta")
    assert r.status_code == 200
    meta = r.get_json()
    assert meta["total_registros"] == 2
    assert len(meta["periodos"]) == 2
    assert "PAMPA ENERGIA" in meta["clientes"] or "SHELL CAPSA" in meta["clientes"]

    r = auth_client.get("/gos/vacaciones/api/tot-hs/resumen")
    assert r.status_code == 200
    resumen = r.get_json()
    assert resumen["registros"] == 2
    assert resumen["total_horas"] == 250
    assert resumen["personas"] == 2


def test_tot_hs_import_real_file_if_present(auth_client, app):
    """Smoke test con el Excel real del usuario si está en Downloads."""
    from pathlib import Path

    path = Path(r"c:\Users\ferna\Downloads\archivo (14).xlsx")
    if not path.is_file():
        return

    with path.open("rb") as f:
        r = auth_client.post(
            "/gos/vacaciones/api/importar/total",
            data={"file": (f, "archivo (14).xlsx")},
            content_type="multipart/form-data",
        )
    assert r.status_code == 200, r.get_json()
    data = r.get_json()
    assert data["detalle"]["registros"] == 679
    assert data["detalle"]["personas"] == 165
    assert data["detalle"]["fecha_min"] == "2025-12-21"
    assert data["detalle"]["fecha_max"] == "2026-07-20"

    r = auth_client.get("/gos/vacaciones/api/tot-hs/por-empleado")
    assert r.status_code == 200
    empleados = r.get_json()
    assert len(empleados) == 165
    assert empleados[0]["total_horas"] > 0


def test_tot_hs_import_rejects_empty_workbook(auth_client):
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "OTRA"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = auth_client.post(
        "/gos/vacaciones/api/importar/total",
        data={"file": (buf, "sin_horas.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 500
    err = r.get_json()["error"].lower()
    assert "tot hs" in err or "período" in err or "periodo" in err or "nombre" in err
