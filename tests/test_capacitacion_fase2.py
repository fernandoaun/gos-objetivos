from gos.extensions import db
from gos.modulos.capacitacion.models import Curso, Participante, Puesto


def test_api_requisitos_por_puesto(auth_client, app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        puesto = Puesto(empresa_id=emp.id, codigo="OP", nombre="Operario")
        curso = Curso(empresa_id=emp.id, codigo="SEG-01", nombre="Seguridad")
        db.session.add_all([puesto, curso])
        db.session.commit()
        puesto_id, curso_id = puesto.id, curso.id

    r = auth_client.post(
        "/gos/capacitacion/api/requisitos",
        json={"puesto_id": puesto_id, "curso_id": curso_id, "obligatorio": True},
    )
    assert r.status_code == 201
    req_id = r.get_json()["requisito"]["id"]

    lista = auth_client.get(f"/gos/capacitacion/api/requisitos?puesto_id={puesto_id}")
    assert len(lista.get_json()["requisitos"]) == 1

    r2 = auth_client.delete(f"/gos/capacitacion/api/requisitos/{req_id}")
    assert r2.status_code == 200


def test_api_actualizar_y_baja_curso(auth_client):
    r = auth_client.post(
        "/gos/capacitacion/api/cursos",
        json={"codigo": "CUR-1", "nombre": "Curso uno", "horas": 4},
    )
    curso_id = r.get_json()["curso"]["id"]

    r2 = auth_client.put(
        f"/gos/capacitacion/api/cursos/{curso_id}",
        json={
            "codigo": "CUR-1",
            "nombre": "Curso actualizado",
            "horas": 8,
            "categoria": "hse",
            "tipo": "obligatoria",
            "origen": "interna",
            "modalidad": "presencial",
        },
    )
    assert r2.status_code == 200
    assert r2.get_json()["curso"]["nombre"] == "Curso actualizado"

    r3 = auth_client.delete(f"/gos/capacitacion/api/cursos/{curso_id}")
    assert r3.status_code == 200

    lista = auth_client.get("/gos/capacitacion/api/cursos")
    assert all(c["id"] != curso_id for c in lista.get_json()["cursos"])


def test_importar_participantes_excel(auth_client, app):
    from io import BytesIO

    import openpyxl

    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        puesto = Puesto(empresa_id=emp.id, codigo="TEC", nombre="Técnico")
        db.session.add(puesto)
        db.session.commit()

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["nombre", "apellido", "legajo", "puesto_codigo"])
    ws2.append(["Ana", "García", "3001", "TEC"])
    buf2 = BytesIO()
    wb2.save(buf2)
    buf2.seek(0)

    r = auth_client.post(
        "/gos/capacitacion/api/participantes/importar",
        data={"archivo": (buf2, "personal.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["creados"] == 1

    lista = auth_client.get("/gos/capacitacion/api/participantes?activo=all")
    nombres = [p["nombre"] for p in lista.get_json()["participantes"]]
    assert any("Ana" in n for n in nombres)


def test_registrar_asistencias(auth_client, app):
    from datetime import date

    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        curso = Curso(empresa_id=emp.id, codigo="IND-01", nombre="Inducción", vigencia_meses=12)
        p = Participante(empresa_id=emp.id, nombre="Pedro", legajo="4001")
        db.session.add_all([curso, p])
        db.session.flush()
        from gos.modulos.capacitacion.models import EncuentroCapacitacion

        enc = EncuentroCapacitacion(
            empresa_id=emp.id,
            curso_id=curso.id,
            titulo="Inducción día 1",
            fecha=date(2026, 3, 10),
            estado="programado",
        )
        db.session.add(enc)
        db.session.commit()
        enc_id, pid = enc.id, p.id

    r = auth_client.put(
        f"/gos/capacitacion/api/encuentros/{enc_id}/cierre",
        json={"personas": [{"participante_id": pid, "asistio": True, "nota": 9}]},
    )
    assert r.status_code == 200
    assert r.get_json()["guardados"] == 1

    det = auth_client.get(f"/gos/capacitacion/api/encuentros/{enc_id}")
    assert det.status_code == 200
    assert det.get_json()["estado"] == "cerrado"


def test_api_programas_por_puesto_y_persona(auth_client, app):
    with app.app_context():
        from gos.models import Empresa
        from gos.modulos.capacitacion.models import InscripcionPrograma, ProgramaCapacitacion

        emp = Empresa.query.first()
        puesto_op = Puesto(empresa_id=emp.id, codigo="OP2", nombre="Operario II")
        puesto_sup = Puesto(empresa_id=emp.id, codigo="SUP2", nombre="Supervisor II")
        persona = Participante(empresa_id=emp.id, nombre="Lucía", legajo="5001", puesto_id=None)
        db.session.add_all([puesto_op, puesto_sup, persona])
        db.session.flush()
        persona.puesto_id = puesto_op.id

        prog_puesto = ProgramaCapacitacion(
            empresa_id=emp.id,
            codigo="PRG-P",
            nombre="Programa operarios",
            alcance="puesto",
            puesto_id=puesto_op.id,
        )
        prog_persona = ProgramaCapacitacion(
            empresa_id=emp.id,
            codigo="PRG-I",
            nombre="Programa individual",
            alcance="persona",
        )
        db.session.add_all([prog_puesto, prog_persona])
        db.session.flush()
        db.session.add(
            InscripcionPrograma(programa_id=prog_persona.id, participante_id=persona.id, estado="inscripto")
        )
        db.session.commit()
        puesto_id, persona_id = puesto_op.id, persona.id

    por_puesto = auth_client.get(f"/gos/capacitacion/api/programas?puesto_id={puesto_id}")
    assert por_puesto.status_code == 200
    codigos_puesto = [p["codigo"] for p in por_puesto.get_json()["programas"]]
    assert "PRG-P" in codigos_puesto
    assert "PRG-I" not in codigos_puesto

    por_persona = auth_client.get(f"/gos/capacitacion/api/programas?participante_id={persona_id}")
    assert por_persona.status_code == 200
    codigos_persona = [p["codigo"] for p in por_persona.get_json()["programas"]]
    assert "PRG-I" in codigos_persona
    assert "PRG-P" in codigos_persona


def _excel_programas(rows):
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programas"
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_importar_programas_excel_crea_y_completa(auth_client, app):
    from gos.modulos.capacitacion.models import PlanCurso, ProgramaCapacitacion, ProgramaPlan, ProgramaPuesto

    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        chofer = Puesto(empresa_id=emp.id, codigo="CHF", nombre="Chofer")
        mecanico = Puesto(empresa_id=emp.id, codigo="MEC", nombre="Mecánico")
        supervisor = Puesto(empresa_id=emp.id, codigo="SUP", nombre="Supervisor")
        curso = Curso(empresa_id=emp.id, codigo="IND-10", nombre="Inducción")
        existente = ProgramaCapacitacion(
            empresa_id=emp.id,
            codigo="LID",
            nombre="Liderazgo",
            tipo="interno",
            alcance="general",
        )
        db.session.add_all([chofer, mecanico, supervisor, curso, existente])
        db.session.flush()
        db.session.add(ProgramaPlan(programa_id=existente.id, nombre="Gestión", orden=1))
        db.session.commit()

    nuevo = _excel_programas(
        [
            ["programa", "codigo", "tipo", "plan", "puesto", "curso"],
            ["Formación Operativa", "FO", "interno", "Seguridad", "Chofer", "Inducción"],
            ["Formación Operativa", "FO", "interno", "Técnico", "Mecánico", ""],
            ["Liderazgo", "", "interno", "Gestión", "Supervisor", ""],
            ["Liderazgo", "", "", "Comunicación", "Jefe de turno", ""],
        ]
    )
    r = auth_client.post(
        "/gos/capacitacion/api/programas/importar",
        data={"archivo": (nuevo, "programas.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["creados"] == 1
    assert data["actualizados"] == 1
    assert data["planes_agregados"] == 3
    assert data["puestos_agregados"] == 3
    assert data["cursos_agregados"] == 1
    assert any("Jefe de turno" in e for e in data["errores"])

    det = auth_client.get("/gos/capacitacion/api/programas?detalle=1")
    programas = {p["nombre"]: p for p in det.get_json()["programas"]}
    fo = programas["Formación Operativa"]
    assert {pl["nombre"] for pl in fo["planes"]} == {"Seguridad", "Técnico"}
    assert {p["nombre"] for p in fo["puestos"]} == {"Chofer", "Mecánico"}
    lid = programas["Liderazgo"]
    assert {pl["nombre"] for pl in lid["planes"]} == {"Gestión", "Comunicación"}
    assert {p["nombre"] for p in lid["puestos"]} == {"Supervisor"}

    with app.app_context():
        fo_db = ProgramaCapacitacion.query.filter_by(codigo="FO").one()
        seg = ProgramaPlan.query.filter_by(programa_id=fo_db.id, nombre="Seguridad").one()
        assert PlanCurso.query.filter_by(plan_id=seg.id).count() == 1
        assert ProgramaPuesto.query.filter_by(programa_id=fo_db.id).count() == 2

    otra_vez = _excel_programas(
        [
            ["programa", "codigo", "tipo", "plan", "puesto", "curso"],
            ["Formación Operativa", "FO", "interno", "Seguridad", "Chofer", "Inducción"],
        ]
    )
    r2 = auth_client.post(
        "/gos/capacitacion/api/programas/importar",
        data={"archivo": (otra_vez, "programas.xlsx")},
        content_type="multipart/form-data",
    )
    assert r2.status_code == 200
    data2 = r2.get_json()
    assert data2["creados"] == 0
    assert data2["planes_agregados"] == 0
    assert data2["puestos_agregados"] == 0
    assert data2["cursos_agregados"] == 0

    plantilla = auth_client.get("/gos/capacitacion/api/programas/importar/plantilla")
    assert plantilla.status_code == 200
    assert "spreadsheet" in plantilla.content_type
    from io import BytesIO

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(plantilla.data))
    headers = [c.value for c in wb.active[1]]
    assert "detalles" in headers
    assert "plan" in headers
    assert "puesto" in headers


def test_importar_programas_excel_completa_detalles(auth_client, app):
    from gos.modulos.capacitacion.models import ProgramaCapacitacion

    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        db.session.add(
            ProgramaCapacitacion(
                empresa_id=emp.id,
                codigo="FO",
                nombre="Formación Operativa",
                tipo="interno",
                descripcion=None,
            )
        )
        db.session.commit()

    buf = _excel_programas(
        [
            ["programa", "detalles"],
            ["Formación Operativa", "Inducción y oficios de planta."],
        ]
    )
    r = auth_client.post(
        "/gos/capacitacion/api/programas/importar",
        data={"archivo": (buf, "programas.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["actualizados"] == 1
    det = auth_client.get("/gos/capacitacion/api/programas?detalle=1")
    fo = next(p for p in det.get_json()["programas"] if p["codigo"] == "FO")
    assert fo["descripcion"] == "Inducción y oficios de planta."


def test_importar_programas_excel_listas_en_una_fila(auth_client, app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        db.session.add_all(
            [
                Puesto(empresa_id=emp.id, codigo="CHF2", nombre="Chofer"),
                Puesto(empresa_id=emp.id, codigo="MEC2", nombre="Mecánico"),
            ]
        )
        db.session.commit()

    buf = _excel_programas(
        [
            ["Programa", "Planes", "Puestos"],
            ["Formación Operativa", "Seguridad; Técnico", "Chofer, Mecánico"],
        ]
    )
    r = auth_client.post(
        "/gos/capacitacion/api/programas/importar",
        data={"archivo": (buf, "programas.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    data = r.get_json()
    assert data["creados"] == 1
    assert data["planes_agregados"] == 2
    assert data["puestos_agregados"] == 2
    assert not data["errores"]

    det = auth_client.get("/gos/capacitacion/api/programas?detalle=1")
    fo = next(p for p in det.get_json()["programas"] if p["nombre"] == "Formación Operativa")
    assert {pl["nombre"] for pl in fo["planes"]} == {"Seguridad", "Técnico"}
    assert {p["nombre"] for p in fo["puestos"]} == {"Chofer", "Mecánico"}
