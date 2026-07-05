"""Tests Fase 7 — pendientes post Fase 6: export, sector puesto, email, CRUD acreditaciones/planes."""

from datetime import date
from io import BytesIO

import openpyxl

from gos.extensions import db
from gos.modulos.capacitacion.models import (
    Acreditacion,
    Curso,
    Participante,
    PlanCapacitacion,
    PlanCurso,
    ProgramaCapacitacion,
    ProgramaPlan,
    ProgramaPuesto,
    Puesto,
)
from gos.modulos.capacitacion.services.export_service import exportar_matriz_analitica_excel


def test_export_matriz_analitica_tabla(app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        buf = exportar_matriz_analitica_excel(emp.id, vista="tabla")
        wb = openpyxl.load_workbook(BytesIO(buf.getvalue()))
        assert wb.active.title == "Matriz tabla"
        assert wb.active.cell(1, 1).value == "Programa"


def test_puesto_con_sector(auth_client, app):
    with app.app_context():
        from gos.models import Empresa
        from gos.modulos.objetivos.models.catalogos import Sector

        emp = Empresa.query.first()
        sector = Sector(empresa_id=emp.id, codigo="S7", nombre="Sector VII")
        db.session.add(sector)
        db.session.commit()
        sector_id = sector.id

    r = auth_client.post(
        "/gos/capacitacion/api/puestos",
        json={"codigo": "P7", "nombre": "Puesto VII", "sector_id": sector_id},
    )
    assert r.status_code == 201
    body = r.get_json()["puesto"]
    assert body["sector_id"] == sector_id
    assert body["sector_nombre"] == "Sector VII"


def test_email_unico_participante(auth_client, app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        db.session.add(
            Participante(empresa_id=emp.id, nombre="Uno", legajo="E7001", email="dup@test.com")
        )
        db.session.commit()

    r = auth_client.post(
        "/gos/capacitacion/api/participantes",
        json={"nombre": "Dos", "legajo": "E7002", "email": "dup@test.com"},
    )
    assert r.status_code == 400
    assert "email" in r.get_json()["error"].lower()


def test_acreditaciones_crud(auth_client, app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        puesto = Puesto(empresa_id=emp.id, codigo="ACR", nombre="Acr")
        curso = Curso(empresa_id=emp.id, codigo="ACR-C", nombre="Curso Acr", horas=2)
        persona = Participante(empresa_id=emp.id, nombre="Acr Persona", legajo="ACR1", puesto_id=None)
        db.session.add_all([puesto, curso, persona])
        db.session.flush()
        persona.puesto_id = puesto.id
        prog = ProgramaCapacitacion(empresa_id=emp.id, codigo="ACR-P", nombre="Prog Acr", tipo="interno")
        db.session.add(prog)
        db.session.flush()
        plan = ProgramaPlan(programa_id=prog.id, nombre="Plan Acr", orden=1)
        db.session.add(plan)
        db.session.flush()
        db.session.add_all([
            ProgramaPuesto(programa_id=prog.id, puesto_id=puesto.id),
            PlanCurso(plan_id=plan.id, curso_id=curso.id, orden=1),
        ])
        db.session.commit()
        ids = {
            "persona_id": persona.id,
            "programa_id": prog.id,
            "plan_id": plan.id,
            "curso_id": curso.id,
        }

    r = auth_client.post(
        "/gos/capacitacion/api/acreditaciones",
        json={**ids, "aprobo": True, "nota": 9},
    )
    assert r.status_code == 201
    acr_id = r.get_json()["acreditacion"]["id"]

    r2 = auth_client.get(f"/gos/capacitacion/api/acreditaciones?persona_id={ids['persona_id']}")
    assert r2.status_code == 200
    assert len(r2.get_json()["acreditaciones"]) == 1

    r3 = auth_client.put(
        f"/gos/capacitacion/api/acreditaciones/{acr_id}",
        json={"nota": 10},
    )
    assert r3.status_code == 200
    assert r3.get_json()["acreditacion"]["nota"] == 10

    r4 = auth_client.delete(f"/gos/capacitacion/api/acreditaciones/{acr_id}")
    assert r4.status_code == 200

    with app.app_context():
        assert Acreditacion.query.get(acr_id) is None


def test_planes_capacitacion_crud(auth_client, app):
    with app.app_context():
        from gos.models import Empresa

        emp = Empresa.query.first()
        curso = Curso(empresa_id=emp.id, codigo="PL-C", nombre="Curso Plan", horas=1)
        persona = Participante(empresa_id=emp.id, nombre="Plan Persona", legajo="PL1")
        db.session.add_all([curso, persona])
        db.session.commit()
        pid, cid = persona.id, curso.id

    r = auth_client.post(
        f"/gos/capacitacion/api/participantes/{pid}/planes-capacitacion",
        json={"curso_id": cid, "fecha_planificada": "2026-08-01"},
    )
    assert r.status_code == 201
    plan_id = r.get_json()["plan"]["id"]
    assert r.get_json()["plan"]["estado"] == "programado"

    r2 = auth_client.get(f"/gos/capacitacion/api/participantes/{pid}/planes-capacitacion")
    assert r2.status_code == 200
    assert len(r2.get_json()["planes"]) == 1

    r3 = auth_client.delete(f"/gos/capacitacion/api/planes-capacitacion/{plan_id}")
    assert r3.status_code == 200

    with app.app_context():
        plan = PlanCapacitacion.query.get(plan_id)
        assert plan.estado == "cancelado"


def test_export_matriz_analitica_endpoint(auth_client):
    r = auth_client.get("/gos/capacitacion/api/matriz/exportar.xlsx?vista=tabla")
    assert r.status_code == 200
    assert "spreadsheetml" in r.content_type
    wb = openpyxl.load_workbook(BytesIO(r.data))
    assert wb.active.title == "Matriz tabla"
