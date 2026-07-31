"""Importación del plan de mantenimiento y reporte mensual desde Excel.

Hojas esperadas:
- Informe: plan anual por unidad con columnas mensuales R / P / E
  - R = referencia / tipo de mantenimiento (1, 2, 3 o 4)
  - P = 1 si ese mantenimiento se programó en el mes
  - E = 1 si ese mantenimiento se ejecutó en el mes
- VTV: unidad + vencimiento VTV
- OTs / Ordenes, TAREAS / Tareas, SOLICITUDES: reporte mensual (Excel fuente Power BI)

Al reimportar se reemplazan por completo los años presentes en cada hoja del archivo.
Los años que no aparecen en el archivo no se tocan.
La hoja SECTOR reemplaza el padrón completo de personal (Alta/Baja).
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from gos.modulos.mantenimiento.models import (
    MantPlanCelda,
    MantPlanMeta,
    MantReporteOrden,
    MantReporteSolicitud,
    MantReporteTarea,
    MantSectorPersona,
    MantUnidad,
    MantVtv,
)

MESES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def normalizar_codigo(raw: str) -> str:
    """HG 01 / HG-01 / hg01 → HG01. También unifica UI↔UL (typo frecuente en el Excel)."""
    s = re.sub(r"[^A-Za-z0-9]", "", (raw or "").strip().upper())
    if s.startswith("UI") and len(s) > 2 and s[2:].isdigit():
        s = "UL" + s[2:]
    return s


def _cell_num(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return 0.0


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Excel serial / Timestamp numérico
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            from datetime import timedelta

            serial = float(value)
            if 20000 <= serial <= 80000:  # ~1954–2119
                return (datetime(1899, 12, 30) + timedelta(days=int(serial))).date()
        except (OverflowError, ValueError):
            pass
    text = str(value).strip()
    if not text or text.lower() in ("nan", "nat", "<na>", "none"):
        return None
    text_date = text[:10] if ("T" in text or " " in text) else text
    for fmt, n in (
        ("%Y-%m-%d", 10),
        ("%d/%m/%Y", 10),
        ("%d-%m-%Y", 10),
        ("%d/%m/%y", 8),
        ("%d-%m-%y", 8),
    ):
        try:
            return datetime.strptime(text_date[:n], fmt).date()
        except ValueError:
            continue
    return None


def _find_sheet(wb, *candidates: str):
    lower = {name.lower().strip(): name for name in wb.sheetnames}
    for cand in candidates:
        if cand.lower() in lower:
            return wb[lower[cand.lower()]]
    for name in wb.sheetnames:
        nl = name.lower()
        if any(c.lower() in nl for c in candidates):
            return wb[name]
    return None


def _get_or_create_unidad(session: Session, nombre: str, cache: dict[str, MantUnidad]) -> MantUnidad:
    codigo = normalizar_codigo(nombre)
    if not codigo:
        raise ValueError(f"Código de unidad inválido: {nombre!r}")
    if codigo in cache:
        unidad = cache[codigo]
        if nombre.strip() and unidad.nombre != nombre.strip():
            # Preferir el nombre con espacios del plan (HG 01) sobre HG-01
            if " " in nombre.strip() or len(nombre.strip()) > len(unidad.nombre):
                unidad.nombre = nombre.strip()
        return unidad

    unidad = session.execute(
        select(MantUnidad).where(MantUnidad.codigo == codigo)
    ).scalar_one_or_none()
    if unidad is None:
        unidad = MantUnidad(codigo=codigo, nombre=nombre.strip() or codigo, activo=True)
        session.add(unidad)
        session.flush()
    else:
        unidad.nombre = nombre.strip() or unidad.nombre
        unidad.activo = True
    cache[codigo] = unidad
    return unidad


def _parse_informe(ws, session: Session, cache: dict[str, MantUnidad]) -> dict:
    titulo = None
    anio = None
    sector = None
    observaciones = None

    for r in range(1, min(9, (ws.max_row or 1) + 1)):
        for c in range(1, min(50, (ws.max_column or 1) + 1)):
            val = ws.cell(r, c).value
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            low = text.lower()
            if titulo is None and "plan de mantenimiento" in low:
                titulo = text
            if low.startswith("fecha:") or (c <= 4 and re.search(r"20\d{2}", text) and "fecha" in low):
                m = re.search(r"(20\d{2})", text)
                if m:
                    anio = int(m.group(1))
            if low == "sector:" or low.startswith("sector"):
                # valor suele estar a la derecha en celdas mergeadas
                for cc in range(c + 1, min(c + 15, (ws.max_column or c) + 1)):
                    vv = ws.cell(r, cc).value
                    if vv and str(vv).strip() and str(vv).strip().lower() not in ("sector:", "sector"):
                        sector = str(vv).strip()
                        break
            if "observacion" in low:
                observaciones = text

    # Año también puede venir solo como número cerca de FECHA
    if anio is None:
        for r in range(1, 9):
            for c in range(1, 10):
                val = ws.cell(r, c).value
                if isinstance(val, int) and 2000 <= val <= 2100:
                    anio = val
                elif isinstance(val, str):
                    m = re.search(r"(20\d{2})", val)
                    if m:
                        anio = int(m.group(1))

    header_row = None
    rpe_row = None
    mes_cols: dict[int, dict[str, int]] = {}  # mes -> {R,P,E: col}

    for r in range(1, min(20, (ws.max_row or 1) + 1)):
        month_hits = []
        for c in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(r, c).value
            if not isinstance(val, str):
                continue
            key = val.strip().lower()
            if key in MESES:
                month_hits.append((c, MESES[key]))
        if len(month_hits) >= 3:
            header_row = r
            rpe_row = r + 1
            for col, mes in month_hits:
                # R P E suelen ocupar col, col+1, col+2
                labels = {}
                for offset, expected in enumerate(("R", "P", "E")):
                    cell_val = ws.cell(rpe_row, col + offset).value
                    label = str(cell_val).strip().upper() if cell_val is not None else expected
                    if label in ("R", "P", "E"):
                        labels[label.lower()] = col + offset
                    else:
                        labels[expected.lower()] = col + offset
                mes_cols[mes] = labels
            break

    if header_row is None or anio is None:
        raise ValueError(
            "No se pudo leer el plan en la hoja Informe "
            "(faltan fila de meses o año FECHA)."
        )

    # Ubicar columna UNIDADES
    unidad_col = 2
    for c in range(1, 8):
        val = ws.cell(header_row, c).value
        if isinstance(val, str) and "unidad" in val.lower():
            unidad_col = c
            break

    celdas = 0
    unidades_plan = 0
    start = rpe_row + 1
    for r in range(start, (ws.max_row or start) + 1):
        raw_nombre = ws.cell(r, unidad_col).value
        if raw_nombre is None or not str(raw_nombre).strip():
            # fila de totales suele tener fórmulas en P/E sin nombre
            continue
        nombre = str(raw_nombre).strip()
        if nombre.lower() in ("total", "totales", "suma"):
            continue
        # ignorar filas sin ningún dato R/P/E
        tiene = False
        for labels in mes_cols.values():
            for col in labels.values():
                if ws.cell(r, col).value not in (None, ""):
                    tiene = True
                    break
            if tiene:
                break
        if not tiene:
            continue

        unidad = _get_or_create_unidad(session, nombre, cache)
        unidades_plan += 1

        # Reemplazar celdas del año para esta unidad
        session.execute(
            delete(MantPlanCelda).where(
                MantPlanCelda.unidad_id == unidad.id,
                MantPlanCelda.anio == anio,
            )
        )

        for mes, labels in mes_cols.items():
            rv = _cell_num(ws.cell(r, labels.get("r", 0)).value) if labels.get("r") else 0.0
            pv = _cell_num(ws.cell(r, labels.get("p", 0)).value) if labels.get("p") else 0.0
            ev = _cell_num(ws.cell(r, labels.get("e", 0)).value) if labels.get("e") else 0.0
            if rv == 0 and pv == 0 and ev == 0:
                continue
            session.add(
                MantPlanCelda(
                    unidad_id=unidad.id,
                    anio=anio,
                    mes=mes,
                    r=rv,
                    p=pv,
                    e=ev,
                )
            )
            celdas += 1

    meta = session.execute(
        select(MantPlanMeta).where(MantPlanMeta.anio == anio)
    ).scalar_one_or_none()
    if meta is None:
        meta = MantPlanMeta(anio=anio)
        session.add(meta)
    meta.titulo = titulo or meta.titulo
    meta.sector = sector or meta.sector
    meta.observaciones = observaciones or meta.observaciones

    return {
        "anio": anio,
        "titulo": titulo,
        "sector": sector,
        "unidades_plan": unidades_plan,
        "celdas": celdas,
    }


def _parse_vtv(ws, session: Session, cache: dict[str, MantUnidad]) -> dict:
    header_row = None
    col_unidad = None
    col_vto = None

    for r in range(1, min(30, (ws.max_row or 1) + 1)):
        for c in range(1, min(20, (ws.max_column or 1) + 1)):
            val = ws.cell(r, c).value
            if not isinstance(val, str):
                continue
            low = val.strip().lower()
            if low in ("unidad", "unidades", "equipo", "vehiculo", "vehículo"):
                header_row = r
                col_unidad = c
            if header_row == r and ("venc" in low or "vtv" in low):
                col_vto = c
        if header_row and col_unidad and col_vto:
            break

    if not header_row or not col_unidad or not col_vto:
        # fallback formato conocido: C=Unidad, D=Vencimiento
        header_row = 5
        col_unidad = 3
        col_vto = 4

    cargados = 0
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        raw = ws.cell(r, col_unidad).value
        if raw is None or not str(raw).strip():
            continue
        nombre = str(raw).strip()
        venc = _parse_date(ws.cell(r, col_vto).value)
        if not venc:
            continue
        unidad = _get_or_create_unidad(session, nombre, cache)
        existing = session.execute(
            select(MantVtv).where(MantVtv.unidad_id == unidad.id)
        ).scalar_one_or_none()
        if existing is None:
            session.add(MantVtv(unidad_id=unidad.id, vencimiento=venc))
        else:
            existing.vencimiento = venc
        cargados += 1

    return {"vtv": cargados}


def _norm_header(value) -> str:
    text = str(value or "").strip().lower()
    text = (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
        .replace("º", "")
        .replace("°", "")
        .replace(".", "")
        .replace("_", " ")
    )
    return re.sub(r"\s+", " ", text).strip()


def _col_map(headers: list) -> dict[str, int]:
    mapping = {}
    for idx, h in enumerate(headers):
        key = _norm_header(h)
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _row_get(row: tuple, cmap: dict[str, int], *aliases: str, default=None):
    for alias in aliases:
        idx = cmap.get(_norm_header(alias))
        if idx is not None and idx < len(row):
            val = row[idx]
            if val is not None and str(val).strip() != "":
                return val
    return default


def _trimestre(mes: int) -> int:
    return ((mes - 1) // 3) + 1


def _str_opt(value, max_len: int) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text[:max_len]


def _nro_str(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        return None
    # "694.0" → "694"
    try:
        f = float(text.replace(",", "."))
        if f.is_integer():
            return str(int(f))
    except ValueError:
        pass
    return text


def _find_header_row(rows: list, *needles_groups: tuple[str, ...]) -> int:
    for i, row in enumerate(rows[:15]):
        joined = " ".join(_norm_header(c) for c in row if c is not None)
        if all(any(n in joined for n in group) for group in needles_groups):
            return i
    return 0


def _replace_by_years(session: Session, model, anios: set[int], rows: list) -> None:
    """Borra solo los años presentes en el archivo e inserta el snapshot completo."""
    if not anios:
        return
    session.execute(delete(model).where(model.anio.in_(sorted(anios))))
    session.flush()
    session.add_all(rows)


def _parse_reporte_ordenes(ws, session: Session) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ordenes": 0, "anios_ordenes": []}
    header_idx = _find_header_row(rows, ("unidad",), ("orden", "estado"))
    cmap = _col_map(list(rows[header_idx]))
    if "unidad" not in cmap:
        raise ValueError("Hoja OTs/Órdenes: no se encontró la columna Unidad.")

    by_nro: dict[str, MantReporteOrden] = {}
    anios: set[int] = set()
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        unidad = str(_row_get(row, cmap, "unidad", "equipo") or "").strip()
        if not unidad:
            continue
        fecha = _parse_date(
            _row_get(row, cmap, "fecha orden", "fecha", "f orden", "alta")
        )
        if not fecha:
            continue
        nro = _nro_str(
            _row_get(row, cmap, "nro orden", "nroorden", "n orden", "orden", "indice")
        ) or f"{unidad}-{fecha.isoformat()}"
        nro_sol = _nro_str(_row_get(row, cmap, "nro solicitud", "nrosolicitud", "solicitud"))
        estado = (
            str(
                _row_get(row, cmap, "estado orden", "estado") or "Sin estado"
            ).strip()
            or "Sin estado"
        )
        by_nro[nro[:64]] = MantReporteOrden(
            nro_orden=nro[:64],
            unidad=unidad[:128],
            estado=estado[:64],
            nro_solicitud=nro_sol[:64] if nro_sol else None,
            estado_solicitud=_str_opt(
                _row_get(row, cmap, "estado solicitud", "estado soli"), 64
            ),
            ingreso_taller=_parse_date(
                _row_get(row, cmap, "ingreso taller", "f ingreso", "ingreso")
            ),
            km=_cell_num(_row_get(row, cmap, "km", default=None))
            if _row_get(row, cmap, "km") is not None
            else None,
            hs=_cell_num(_row_get(row, cmap, "hs", "horas", default=None))
            if _row_get(row, cmap, "hs", "horas") is not None
            else None,
            fecha=fecha,
            anio=fecha.year,
            mes=fecha.month,
            trimestre=_trimestre(fecha.month),
        )
        anios.add(fecha.year)

    parsed = list(by_nro.values())
    _replace_by_years(session, MantReporteOrden, anios, parsed)
    return {"ordenes": len(parsed), "anios_ordenes": sorted(anios)}


def _parse_reporte_tareas(ws, session: Session) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"tareas": 0, "anios_tareas": []}
    header_idx = _find_header_row(rows, ("unidad",), ("hora", "clase", "tarea"))
    cmap = _col_map(list(rows[header_idx]))
    if "unidad" not in cmap:
        raise ValueError("Hoja Tareas: no se encontró la columna Unidad.")

    by_nro: dict[str, MantReporteTarea] = {}
    anios: set[int] = set()
    for idx, row in enumerate(rows[header_idx + 1 :]):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        unidad = str(_row_get(row, cmap, "unidad", "equipo") or "").strip()
        if not unidad:
            continue
        fecha = _parse_date(_row_get(row, cmap, "fecha", "fecha tarea", "alta"))
        if not fecha:
            continue
        horas = _cell_num(
            _row_get(row, cmap, "total horas", "horas", "tiempo neto", default=0)
        )
        nro_orden = _nro_str(_row_get(row, cmap, "orden", "nro orden", "nroorden"))
        nro_tarea = _nro_str(_row_get(row, cmap, "tarea", "nro tarea", "nrotarea"))
        if not nro_tarea:
            nro_tarea = f"{nro_orden or unidad}-{fecha.isoformat()}-{idx}"
        desc = _str_opt(_row_get(row, cmap, "descripcion", "descripción"), 500)
        by_nro[nro_tarea[:64]] = MantReporteTarea(
            nro_tarea=nro_tarea[:64],
            nro_orden=nro_orden[:64] if nro_orden else None,
            unidad=unidad[:128],
            tipo=_str_opt(_row_get(row, cmap, "tipo"), 64),
            clase=_str_opt(_row_get(row, cmap, "clase"), 64),
            categoria=_str_opt(_row_get(row, cmap, "categoria"), 64),
            lugar=_str_opt(_row_get(row, cmap, "lugar"), 64),
            estado=_str_opt(_row_get(row, cmap, "estado"), 64),
            solicitante=_str_opt(_row_get(row, cmap, "solicitante"), 128),
            urgencia=_str_opt(_row_get(row, cmap, "urgencia"), 64),
            descripcion=desc,
            cant_personal=_cell_num(
                _row_get(
                    row,
                    cmap,
                    "cant personal",
                    "cantidad personal",
                    "cant. personal",
                    "cantpersonal",
                    "personal",
                    default=None,
                )
            )
            if _row_get(
                row,
                cmap,
                "cant personal",
                "cantidad personal",
                "cant. personal",
                "cantpersonal",
                "personal",
            )
            is not None
            else None,
            tercerizado=_str_opt(_row_get(row, cmap, "tercerizado"), 64),
            total_horas=horas,
            fecha=fecha,
            anio=fecha.year,
            mes=fecha.month,
            trimestre=_trimestre(fecha.month),
        )
        anios.add(fecha.year)

    parsed = list(by_nro.values())
    _replace_by_years(session, MantReporteTarea, anios, parsed)
    return {"tareas": len(parsed), "anios_tareas": sorted(anios)}


def _parse_reporte_solicitudes(ws, session: Session) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"solicitudes": 0, "anios_solicitudes": []}
    header_idx = _find_header_row(rows, ("unidad",), ("solicitud", "soli"))
    cmap = _col_map(list(rows[header_idx]))
    if "unidad" not in cmap:
        raise ValueError("Hoja Solicitudes: no se encontró la columna Unidad.")

    by_nro: dict[str, MantReporteSolicitud] = {}
    anios: set[int] = set()
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        unidad = str(_row_get(row, cmap, "unidad", "equipo") or "").strip()
        if not unidad:
            continue
        fecha = _parse_date(_row_get(row, cmap, "fecha", "fecha solicitud", "alta"))
        if not fecha:
            continue
        nro = _nro_str(
            _row_get(row, cmap, "nro solicitud", "nrosolicitud", "solicitud", "indice")
        ) or f"{unidad}-{fecha.isoformat()}"
        nro_orden = _nro_str(_row_get(row, cmap, "nro orden", "nroorden", "orden"))
        # 0 en Excel = sin orden asociada
        if nro_orden in ("0", "0.0"):
            nro_orden = None
        estado = (
            str(
                _row_get(row, cmap, "estado soli", "estado solicitud", "estado")
                or "Sin estado"
            ).strip()
            or "Sin estado"
        )
        by_nro[nro[:64]] = MantReporteSolicitud(
            nro_solicitud=nro[:64],
            fecha=fecha,
            unidad=unidad[:128],
            tipo=_str_opt(_row_get(row, cmap, "tipo"), 128),
            estado=estado[:64],
            solicitante=_str_opt(_row_get(row, cmap, "solicitante"), 128),
            fecha_solicitud_ingreso=_parse_date(
                _row_get(
                    row,
                    cmap,
                    "fsolicingresso",
                    "fsolic ingreso",
                    "fecha solicitud ingreso",
                )
            ),
            fecha_ingresar=_parse_date(
                _row_get(row, cmap, "fingresar", "f ingresar", "fecha ingresar")
            ),
            fecha_retirar=_parse_date(
                _row_get(row, cmap, "fretirar", "f retirar", "fecha retirar")
            ),
            nro_orden=nro_orden[:64] if nro_orden else None,
            estado_orden=_str_opt(_row_get(row, cmap, "estado orden"), 64),
            anio=fecha.year,
            mes=fecha.month,
            trimestre=_trimestre(fecha.month),
        )
        anios.add(fecha.year)

    parsed = list(by_nro.values())
    _replace_by_years(session, MantReporteSolicitud, anios, parsed)
    return {"solicitudes": len(parsed), "anios_solicitudes": sorted(anios)}


def _parse_sector_personas(ws, session: Session) -> dict:
    """Importa personal del sector (Alta / Baja). Sin baja = sigue activo."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"sector_personas": 0}
    header_idx = _find_header_row(rows, ("legajo", "nombre"), ("alta",))
    cmap = _col_map(list(rows[header_idx]))
    if "legajo" not in cmap and "nombre" not in cmap:
        raise ValueError("Hoja SECTOR: no se encontró Legajo/Nombre.")

    by_legajo: dict[str, MantSectorPersona] = {}
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        legajo = _nro_str(_row_get(row, cmap, "legajo"))
        nombre = str(_row_get(row, cmap, "nombre", "apellido y nombre") or "").strip()
        alta = _parse_date(_row_get(row, cmap, "alta", "fecha alta", "f alta"))
        if not alta or not (legajo or nombre):
            continue
        if not legajo:
            legajo = nombre[:64]
        baja = _parse_date(_row_get(row, cmap, "baja", "fecha baja", "f baja"))
        by_legajo[legajo[:64]] = MantSectorPersona(
            legajo=legajo[:64],
            nombre=(nombre or legajo)[:255],
            fecha_alta=alta,
            fecha_baja=baja,
            localidad_real=_str_opt(
                _row_get(
                    row,
                    cmap,
                    "localidad real",
                    "localidad",
                    "domicilio",
                ),
                128,
            ),
            funcion_general=_str_opt(
                _row_get(row, cmap, "funcion general", "función general"), 128
            ),
            funcion=_str_opt(_row_get(row, cmap, "funcion", "función"), 128),
            grupo=_str_opt(_row_get(row, cmap, "grupo"), 64),
            turno=_str_opt(_row_get(row, cmap, "turno"), 64),
        )

    parsed = list(by_legajo.values())
    session.execute(delete(MantSectorPersona))
    session.add_all(parsed)
    return {"sector_personas": len(parsed)}


def _find_ordenes_sheet(wb):
    """Prioriza OTs / Ordenes; evita coincidir con SOLICITUDES."""
    lower = {name.lower().strip(): name for name in wb.sheetnames}
    for cand in ("ots", "ordenes", "órdenes", "orden"):
        if cand in lower:
            return wb[lower[cand]]
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if "solicitud" in nl:
            continue
        if nl in ("ots", "ot") or "orden" in nl:
            return wb[name]
    return None


def import_vtv_excel(path: str | Path, session: Session) -> dict:
    path = Path(path)
    # read_only ahorra memoria en Excels grandes (TAREAS ~10k+ filas).
    # Informe/VTV usan ws.cell → necesitan workbook normal.
    wb_peek = load_workbook(path, read_only=True, data_only=True)
    sheetnames = list(wb_peek.sheetnames)
    has_informe = _find_sheet(wb_peek, "Informe", "Plan", "Mantenimiento") is not None
    has_vtv = _find_sheet(wb_peek, "VTV", "Vtv") is not None
    wb_peek.close()

    if has_informe or has_vtv:
        wb = load_workbook(path, data_only=True)
    else:
        wb = load_workbook(path, read_only=True, data_only=True)

    cache: dict[str, MantUnidad] = {}

    informe = _find_sheet(wb, "Informe", "Plan", "Mantenimiento")
    vtv_sheet = _find_sheet(wb, "VTV", "Vtv")
    ordenes_sheet = _find_ordenes_sheet(wb)
    tareas_sheet = _find_sheet(wb, "TAREAS", "Tareas", "Tarea")
    solicitudes_sheet = _find_sheet(wb, "SOLICITUDES", "Solicitudes", "Solicitud")
    sector_sheet = _find_sheet(wb, "SECTOR", "Sector", "Personal")

    result = {
        "anio": None,
        "titulo": None,
        "sector": None,
        "unidades_plan": 0,
        "celdas": 0,
        "vtv": 0,
        "ordenes": 0,
        "tareas": 0,
        "solicitudes": 0,
        "sector_personas": 0,
        "anios_ordenes": [],
        "anios_tareas": [],
        "anios_solicitudes": [],
        "hojas": sheetnames,
    }

    if (
        informe is None
        and vtv_sheet is None
        and ordenes_sheet is None
        and tareas_sheet is None
        and solicitudes_sheet is None
        and sector_sheet is None
    ):
        wb.close()
        raise ValueError(
            "El Excel no tiene hojas 'Informe', 'VTV', 'OTs'/'Ordenes', 'Tareas', "
            "'Solicitudes' ni 'SECTOR'. "
            f"Hojas encontradas: {', '.join(sheetnames)}"
        )

    if informe is not None:
        info = _parse_informe(informe, session, cache)
        result.update(info)

    if vtv_sheet is not None:
        vtv_info = _parse_vtv(vtv_sheet, session, cache)
        result.update(vtv_info)

    if ordenes_sheet is not None:
        result.update(_parse_reporte_ordenes(ordenes_sheet, session))

    if tareas_sheet is not None:
        result.update(_parse_reporte_tareas(tareas_sheet, session))

    if solicitudes_sheet is not None:
        result.update(_parse_reporte_solicitudes(solicitudes_sheet, session))

    if sector_sheet is not None:
        result.update(_parse_sector_personas(sector_sheet, session))

    session.commit()
    wb.close()
    return result
