from __future__ import annotations

import re
import unicodedata
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from gos.extensions import db
from gos.modulos.capacitacion.models import (
    Centro,
    Curso,
    Participante,
    PlanCurso,
    ProgramaCapacitacion,
    ProgramaPlan,
    ProgramaPuesto,
    Puesto,
)
from gos.modulos.capacitacion.models.programa import TIPOS_PROGRAMA
from gos.modulos.capacitacion.services.taxonomia_service import (
    clasificacion_desde_legacy,
    tipo_capacitacion_legacy,
    validar_clasificacion,
)
from gos.modulos.capacitacion.services.catalogo_service import (
    _parse_decimal,
    _parse_int,
    centro_id_desde_texto,
    puesto_id_desde_texto,
    sector_id_desde_texto,
)
from gos.modulos.capacitacion.services.programa_service import _ensure_requisito


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(ws) -> dict[str, int]:
    headers = {}
    for col, cell in enumerate(ws[1], start=1):
        key = _cell_str(cell.value).lower().replace(" ", "_")
        if key:
            headers[key] = col
    return headers


def importar_participantes_excel(empresa_id: int, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    required = {"nombre", "legajo"}
    if not required.issubset(headers):
        raise ValueError(
            "El Excel debe tener encabezados en la fila 1. Mínimo: nombre, legajo. "
            "Opcionales: apellido, email, centro, centro_codigo, sector, sector_codigo, "
            "puesto, puesto_codigo, observaciones. "
            "Si la persona ya existe (mismo legajo), solo se actualizan puesto, centro y sector."
        )

    legajos_existentes = {
        p.legajo for p in Participante.query.filter_by(empresa_id=empresa_id).all() if p.legajo
    }

    creados = actualizados = omitidos = 0
    errores: list[str] = []
    puestos_cambiaron = False

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def val(key: str):
            col = headers.get(key)
            if not col:
                return ""
            return _cell_str(row[col - 1])

        def val_any(*keys: str) -> str:
            for key in keys:
                value = val(key)
                if value:
                    return value
            return ""

        nombre = val("nombre")
        if not nombre:
            omitidos += 1
            continue

        legajo = val("legajo") or None
        if not legajo:
            errores.append(f"Fila {row_idx}: el legajo es obligatorio")
            continue

        sector_texto = val_any("sector_codigo", "sector")
        puesto_texto = val_any("puesto_codigo", "puesto")
        centro_codigo = val("centro_codigo")
        centro_texto = val_any("centro")

        sector_id = sector_id_desde_texto(empresa_id, sector_texto) if sector_texto else None
        puesto_id = (
            puesto_id_desde_texto(empresa_id, puesto_texto, sector_id=sector_id)
            if puesto_texto
            else None
        )
        centro_id = None
        if centro_codigo:
            centro = Centro.query.filter_by(
                empresa_id=empresa_id, codigo=centro_codigo, activo=True
            ).first()
            if not centro:
                errores.append(f"Fila {row_idx}: centro «{centro_codigo}» no encontrado")
                continue
            centro_id = centro.id
        elif centro_texto:
            centro_id = centro_id_desde_texto(empresa_id, centro_texto)

        existente = None
        if legajo:
            existente = Participante.query.filter_by(empresa_id=empresa_id, legajo=legajo).first()

        if existente:
            # Re-import: solo actualiza puesto / centro / sector (no pisa el resto).
            changed = False
            puesto_cambio = False
            if puesto_id is not None and existente.puesto_id != puesto_id:
                existente.puesto_id = puesto_id
                changed = True
                puesto_cambio = True
                puestos_cambiaron = True
            if centro_id is not None and existente.centro_id != centro_id:
                existente.centro_id = centro_id
                changed = True
            if sector_id is not None and existente.sector_id != sector_id:
                existente.sector_id = sector_id
                changed = True
            elif puesto_cambio and sector_id is None:
                # Si el Excel trae puesto pero no sector, alinear sector del puesto.
                from gos.modulos.capacitacion.models import Puesto

                puesto = Puesto.query.filter_by(id=puesto_id, empresa_id=empresa_id).first()
                if puesto and puesto.sector_id and existente.sector_id != puesto.sector_id:
                    existente.sector_id = puesto.sector_id
                    changed = True
            if changed:
                actualizados += 1
            else:
                omitidos += 1
            continue

        data = {
            "nombre": nombre,
            "apellido": val("apellido") or None,
            "legajo": legajo,
            "email": val("email") or None,
            "centro_id": centro_id,
            "observaciones": val("observaciones") or None,
            "sector_id": sector_id,
            "puesto_id": puesto_id,
        }

        if legajo and legajo in legajos_existentes:
            errores.append(f"Fila {row_idx}: legajo duplicado «{legajo}»")
            continue
        from gos.modulos.capacitacion.services.catalogo_service import crear_participante

        try:
            crear_participante(empresa_id, data)
        except ValueError as exc:
            errores.append(f"Fila {row_idx}: {exc}")
            continue
        if legajo:
            legajos_existentes.add(legajo)
        creados += 1

    db.session.commit()
    if puestos_cambiaron:
        from gos.modulos.capacitacion.services.acreditacion_service import refrescar_vigencias
        from gos.modulos.capacitacion.services.catalogo_service import desactivar_puestos_huerfanos

        desactivar_puestos_huerfanos(empresa_id, gracia_horas=None)
        refrescar_vigencias(empresa_id)
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores}


def importar_cursos_excel(empresa_id: int, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    if "codigo" not in headers or "nombre" not in headers:
        raise ValueError(
            "El Excel debe tener encabezados: codigo, nombre. "
            "Opcionales: descripcion, categoria, tipo, origen, modalidad, horas, "
            "vigencia_meses, requiere_evaluacion, puntaje_minimo. "
            "Legado: tipo_capacitacion (se mapea a la cascada)"
        )

    creados = actualizados = omitidos = 0
    errores: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def val(key: str):
            col = headers.get(key)
            if not col:
                return ""
            return _cell_str(row[col - 1])

        codigo = val("codigo")
        nombre = val("nombre")
        if not codigo or not nombre:
            omitidos += 1
            continue

        modalidad = val("modalidad").lower() or None
        categoria = val("categoria").lower() or None
        tipo = val("tipo").lower() or None
        origen = val("origen").lower() or None
        legacy_tipo = val("tipo_capacitacion").lower() or None

        if not categoria and legacy_tipo:
            categoria, tipo, origen = clasificacion_desde_legacy(empresa_id, legacy_tipo)

        try:
            categoria, tipo, origen, modalidad = validar_clasificacion(
                empresa_id, categoria, tipo, origen, modalidad
            )
        except ValueError as exc:
            errores.append(f"Fila {row_idx}: {exc}")
            continue

        try:
            horas = _parse_decimal(val("horas") or None)
            vigencia = _parse_int(val("vigencia_meses") or None)
            puntaje = _parse_decimal(val("puntaje_minimo") or None)
        except ValueError as exc:
            errores.append(f"Fila {row_idx}: {exc}")
            continue

        requiere_eval = val("requiere_evaluacion").lower() in ("1", "true", "si", "sí", "yes")

        curso = Curso.query.filter_by(empresa_id=empresa_id, codigo=codigo).first()
        if curso:
            curso.nombre = nombre
            curso.descripcion = val("descripcion") or None
            curso.categoria = categoria
            curso.tipo = tipo
            curso.origen = origen
            curso.tipo_capacitacion = tipo_capacitacion_legacy(categoria, tipo)
            curso.horas = horas
            curso.modalidad = modalidad
            curso.vigencia_meses = vigencia
            curso.requiere_evaluacion = requiere_eval
            curso.puntaje_minimo = puntaje
            curso.activo = True
            actualizados += 1
        else:
            db.session.add(
                Curso(
                    empresa_id=empresa_id,
                    codigo=codigo,
                    nombre=nombre,
                    descripcion=val("descripcion") or None,
                    categoria=categoria,
                    tipo=tipo,
                    origen=origen,
                    tipo_capacitacion=tipo_capacitacion_legacy(categoria, tipo),
                    horas=horas,
                    modalidad=modalidad,
                    vigencia_meses=vigencia,
                    requiere_evaluacion=requiere_eval,
                    puntaje_minimo=puntaje,
                )
            )
            creados += 1

    db.session.commit()
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos, "errores": errores}


_HEADER_PROGRAMA = {
    "programa": {"programa", "nombre", "programa_nombre", "nombre_programa"},
    "codigo": {"codigo", "programa_codigo", "cod", "codigo_programa"},
    "tipo": {"tipo", "tipo_programa"},
    "descripcion": {"descripcion", "observaciones"},
    "plan": {"plan", "planes", "nombre_plan", "plan_nombre"},
    "puesto": {
        "puesto",
        "puestos",
        "puesto_nombre",
        "puestos_que_aplican",
        "puesto_codigo",
        "puestos_codigo",
    },
    "curso": {"curso", "cursos", "curso_nombre", "nombre_curso"},
    "curso_codigo": {"curso_codigo", "codigo_curso"},
}

_SPLIT_VALORES = re.compile(r"[\n;|]+")


def _fold(texto: str) -> str:
    s = unicodedata.normalize("NFD", texto or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return " ".join(s.lower().replace("_", " ").split())


def _header_map_programas(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col, cell in enumerate(ws[1], start=1):
        folded = _fold(_cell_str(cell.value))
        key = folded.replace(" ", "_")
        if not key:
            continue
        for canon, aliases in _HEADER_PROGRAMA.items():
            if canon in headers:
                continue
            if key in aliases or folded in aliases:
                headers[canon] = col
                break
    return headers


def _split_valores(raw: str) -> list[str]:
    if not raw:
        return []
    partes: list[str] = []
    vistos: set[str] = set()
    for chunk in _SPLIT_VALORES.split(raw):
        chunk = chunk.strip()
        if not chunk:
            continue
        items = [p.strip() for p in chunk.split(",") if p.strip()] if "," in chunk else [chunk]
        for item in items:
            clave = _fold(item)
            if not clave or clave in vistos:
                continue
            vistos.add(clave)
            partes.append(item)
    return partes


def _val_fila(headers: dict[str, int], row, *keys: str) -> str:
    for key in keys:
        col = headers.get(key)
        if not col:
            continue
        value = _cell_str(row[col - 1])
        if value:
            return value
    return ""


def _indexar_por_texto(items, *attrs: str) -> dict[str, object]:
    idx: dict[str, object] = {}
    for item in items:
        for attr in attrs:
            clave = _fold(str(getattr(item, attr, "") or ""))
            if clave and clave not in idx:
                idx[clave] = item
    return idx


def _generar_codigo_programa(empresa_id: int, nombre: str, usados: set[str]) -> str:
    base = "".join(ch for ch in nombre.upper() if ch.isalnum())[:12] or "PROG"
    codigo = base
    n = 1
    while (
        codigo.lower() in usados
        or ProgramaCapacitacion.query.filter_by(empresa_id=empresa_id, codigo=codigo).first()
    ):
        codigo = f"{base}{n}"
        n += 1
    usados.add(codigo.lower())
    return codigo


def _parse_tipo_programa(raw: str) -> str | None:
    tipo = _fold(raw)
    if not tipo:
        return None
    if tipo in ("interno", "interna", "i"):
        return "interno"
    if tipo in ("externo", "externa", "e"):
        return "externo"
    if tipo in TIPOS_PROGRAMA:
        return tipo
    return None


def _resolver_programa(
    programas_codigo: dict[str, ProgramaCapacitacion],
    programas_nombre: dict[str, list[ProgramaCapacitacion]],
    codigo: str,
    nombre: str,
    fila: int,
    errores: list[str],
) -> ProgramaCapacitacion | None | str:
    """Devuelve el programa, None si hay que crearlo, o 'error' si es ambiguo."""
    if codigo:
        hallado = programas_codigo.get(_fold(codigo))
        if hallado:
            return hallado
    if nombre:
        candidatos = programas_nombre.get(_fold(nombre)) or []
        if len(candidatos) == 1:
            return candidatos[0]
        if len(candidatos) > 1:
            errores.append(
                f"Fila {fila}: hay varios programas llamados «{nombre}». "
                "Indicá el código para saber cuál completar."
            )
            return "error"
    return None


def _agregar_plan_si_falta(programa: ProgramaCapacitacion, nombre: str) -> tuple[ProgramaPlan, bool]:
    clave = _fold(nombre)
    for plan in programa.planes.all():
        if _fold(plan.nombre) == clave:
            return plan, False
    max_orden = (
        db.session.query(db.func.max(ProgramaPlan.orden)).filter_by(programa_id=programa.id).scalar()
        or 0
    )
    plan = ProgramaPlan(programa_id=programa.id, nombre=nombre.strip(), orden=max_orden + 1)
    db.session.add(plan)
    db.session.flush()
    return plan, True


def _agregar_puesto_si_falta(programa: ProgramaCapacitacion, puesto_id: int) -> bool:
    if ProgramaPuesto.query.filter_by(programa_id=programa.id, puesto_id=puesto_id).first():
        return False
    db.session.add(ProgramaPuesto(programa_id=programa.id, puesto_id=puesto_id))
    db.session.flush()
    if not programa.puesto_id:
        programa.puesto_id = puesto_id
    programa.alcance = "puesto"
    return True


def _agregar_curso_si_falta(
    empresa_id: int, programa: ProgramaCapacitacion, plan: ProgramaPlan, curso: Curso
) -> bool:
    if PlanCurso.query.filter_by(plan_id=plan.id, curso_id=curso.id).first():
        return False
    max_orden = db.session.query(db.func.max(PlanCurso.orden)).filter_by(plan_id=plan.id).scalar() or 0
    db.session.add(PlanCurso(plan_id=plan.id, curso_id=curso.id, orden=max_orden + 1))
    db.session.flush()
    for pp in programa.puestos_asignados.all():
        _ensure_requisito(empresa_id, pp.puesto_id, curso.id)
    if programa.puesto_id:
        _ensure_requisito(empresa_id, programa.puesto_id, curso.id)
    return True


def _cursos_del_programa(programa: ProgramaCapacitacion) -> list[int]:
    ids: list[int] = []
    vistos: set[int] = set()
    for plan in programa.planes.all():
        for pc in plan.cursos.all():
            if pc.curso_id and pc.curso_id not in vistos:
                vistos.add(pc.curso_id)
                ids.append(pc.curso_id)
    return ids


def _fusionar_grupo(destino: dict, origen: dict) -> None:
    if origen["nombre"] and not destino["nombre"]:
        destino["nombre"] = origen["nombre"]
    if origen["codigo"] and not destino["codigo"]:
        destino["codigo"] = origen["codigo"]
    if destino["tipo"] is None:
        destino["tipo"] = origen["tipo"]
    if not destino["descripcion"]:
        destino["descripcion"] = origen["descripcion"]
    destino["filas"].extend(origen["filas"])
    for key, puesto_txt in origen["puestos"].items():
        destino["puestos"].setdefault(key, puesto_txt)
    for plan_key, plan_data in origen["planes"].items():
        dest_plan = destino["planes"].setdefault(
            plan_key, {"nombre": plan_data["nombre"], "cursos": []}
        )
        conocidos = {_fold(c) for c in dest_plan["cursos"]}
        for curso_txt in plan_data["cursos"]:
            if _fold(curso_txt) not in conocidos:
                dest_plan["cursos"].append(curso_txt)
                conocidos.add(_fold(curso_txt))


def _fusionar_grupos_programa(
    grupos: dict[str, dict], orden: list[str]
) -> tuple[dict[str, dict], list[str]]:
    """Une filas del mismo programa cuando unas traen código y otras solo el nombre."""
    por_nombre: dict[str, list[str]] = {}
    for clave, grupo in grupos.items():
        nombre = _fold(grupo.get("nombre") or "")
        if nombre:
            por_nombre.setdefault(nombre, []).append(clave)

    absorbidos: set[str] = set()
    for clave in list(orden):
        if not clave.startswith("n:"):
            continue
        grupo = grupos.get(clave)
        if not grupo:
            continue
        nombre = _fold(grupo.get("nombre") or "")
        candidatos = [
            k
            for k in por_nombre.get(nombre, [])
            if k.startswith("c:") and k not in absorbidos
        ]
        if len(candidatos) == 1:
            _fusionar_grupo(grupos[candidatos[0]], grupo)
            absorbidos.add(clave)

    nuevo_orden = [k for k in orden if k not in absorbidos]
    nuevo_grupos = {k: grupos[k] for k in nuevo_orden}
    return nuevo_grupos, nuevo_orden


def plantilla_programas_excel() -> BytesIO:
    """Excel de ejemplo para cargar programas, planes y puestos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programas"

    headers = ["programa", "codigo", "tipo", "plan", "puesto", "curso"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(1, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ejemplos = [
        ["Formación Operativa", "FO", "interno", "Seguridad", "Chofer", "Inducción"],
        ["Formación Operativa", "FO", "interno", "Seguridad", "Chofer", "Primeros auxilios"],
        ["Formación Operativa", "FO", "interno", "Técnico", "Mecánico", ""],
        ["Liderazgo", "", "interno", "Gestión", "Supervisor, Jefe de turno", ""],
    ]
    for row in ejemplos:
        ws.append(row)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28

    tipo_dv = DataValidation(type="list", formula1='"interno,externo"', allow_blank=True)
    tipo_dv.error = "Usá interno o externo"
    tipo_dv.errorTitle = "Tipo inválido"
    ws.add_data_validation(tipo_dv)
    tipo_dv.add("C2:C500")

    inst = wb.create_sheet("Instrucciones")
    inst["A1"] = "Cómo importar programas"
    inst["A1"].font = Font(bold=True, size=14)
    lineas = [
        "",
        "Usá este archivo (o uno propio con los mismos encabezados) en Capacitación → Programas → Importar Excel.",
        "",
        "Columnas:",
        "• programa (obligatorio): nombre del programa. Si ya existe, se completa; si no, se crea.",
        "• codigo: opcional. Sirve para identificar un programa ya cargado. Si lo dejás vacío se genera solo.",
        "• tipo: interno o externo. Si falta, se asume interno.",
        "• plan: nombre del plan. Podés repetir filas (un plan por fila) o separar varios con coma o punto y coma.",
        "• puesto: nombre o código del puesto ya cargado en el catálogo. También admite varios separados por coma.",
        "• curso: opcional. Nombre o código de un curso ya cargado; se agrega al plan de esa fila.",
        "",
        "Comportamiento:",
        "• Programas nuevos: se crean con los planes y puestos del Excel.",
        "• Programas que ya existen: solo se agregan los planes y puestos que falten. No se borra nada.",
        "• Los puestos y cursos tienen que existir en el sistema; si un nombre no coincide, esa fila se informa como error.",
        "• Podés volver a importar el mismo archivo: lo que ya esté no se duplica.",
    ]
    for i, linea in enumerate(lineas, start=1):
        inst.cell(i, 1, linea)
    inst.column_dimensions["A"].width = 120

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def importar_programas_excel(empresa_id: int, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = None
    for name in wb.sheetnames:
        if _fold(name) in {"programas", "programa", "datos"}:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    headers = _header_map_programas(ws)
    if "programa" not in headers and "codigo" not in headers:
        raise ValueError(
            "El Excel debe tener encabezados en la fila 1. Mínimo: programa. "
            "Opcionales: codigo, tipo, plan, puesto, curso. "
            "Descargá la plantilla desde Programas para ver el formato."
        )

    programas = ProgramaCapacitacion.query.filter_by(empresa_id=empresa_id).all()
    programas_codigo = _indexar_por_texto(programas, "codigo")
    programas_nombre: dict[str, list[ProgramaCapacitacion]] = {}
    for prog in programas:
        programas_nombre.setdefault(_fold(prog.nombre), []).append(prog)

    puestos_idx = _indexar_por_texto(
        Puesto.query.filter_by(empresa_id=empresa_id, activo=True).all(), "codigo", "nombre"
    )
    cursos_idx = _indexar_por_texto(
        Curso.query.filter_by(empresa_id=empresa_id, activo=True).all(), "codigo", "nombre"
    )

    grupos: dict[str, dict] = {}
    orden_grupos: list[str] = []
    errores: list[str] = []
    omitidos = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        nombre = _val_fila(headers, row, "programa")
        codigo = _val_fila(headers, row, "codigo")
        if not nombre and not codigo:
            omitidos += 1
            continue

        clave = f"c:{_fold(codigo)}" if codigo else f"n:{_fold(nombre)}"
        grupo = grupos.get(clave)
        if grupo is None:
            grupo = {
                "nombre": nombre,
                "codigo": codigo,
                "tipo": _parse_tipo_programa(_val_fila(headers, row, "tipo")),
                "descripcion": _val_fila(headers, row, "descripcion"),
                "planes": {},
                "puestos": {},
                "filas": [],
            }
            grupos[clave] = grupo
            orden_grupos.append(clave)
        else:
            if nombre and not grupo["nombre"]:
                grupo["nombre"] = nombre
            if codigo and not grupo["codigo"]:
                grupo["codigo"] = codigo
            if grupo["tipo"] is None:
                grupo["tipo"] = _parse_tipo_programa(_val_fila(headers, row, "tipo"))
            if not grupo["descripcion"]:
                grupo["descripcion"] = _val_fila(headers, row, "descripcion")
        grupo["filas"].append(row_idx)

        tipo_raw = _val_fila(headers, row, "tipo")
        if tipo_raw and _parse_tipo_programa(tipo_raw) is None:
            errores.append(f"Fila {row_idx}: tipo «{tipo_raw}» inválido (interno/externo)")

        planes_fila = _split_valores(_val_fila(headers, row, "plan"))
        cursos_fila = _split_valores(_val_fila(headers, row, "curso"))
        cursos_fila.extend(_split_valores(_val_fila(headers, row, "curso_codigo")))
        for plan_nombre in planes_fila:
            plan_key = _fold(plan_nombre)
            plan_data = grupo["planes"].setdefault(plan_key, {"nombre": plan_nombre, "cursos": []})
            for curso_txt in cursos_fila:
                if _fold(curso_txt) not in {_fold(c) for c in plan_data["cursos"]}:
                    plan_data["cursos"].append(curso_txt)
        if cursos_fila and not planes_fila:
            errores.append(
                f"Fila {row_idx}: el curso «{cursos_fila[0]}» no se asoció porque falta el plan"
            )

        for puesto_txt in _split_valores(_val_fila(headers, row, "puesto")):
            grupo["puestos"].setdefault(_fold(puesto_txt), puesto_txt)

    grupos, orden_grupos = _fusionar_grupos_programa(grupos, orden_grupos)

    creados = actualizados = 0
    planes_agregados = puestos_agregados = cursos_agregados = 0
    codigos_usados = {(_fold(p.codigo) if p.codigo else "") for p in programas}

    for clave in orden_grupos:
        grupo = grupos[clave]
        fila = grupo["filas"][0]
        nombre = (grupo["nombre"] or "").strip()
        codigo = (grupo["codigo"] or "").strip()
        if not nombre and not codigo:
            omitidos += 1
            continue

        hallado = _resolver_programa(
            programas_codigo, programas_nombre, codigo, nombre, fila, errores
        )
        if hallado == "error":
            continue

        es_nuevo = hallado is None
        if es_nuevo:
            if not nombre:
                errores.append(f"Fila {fila}: para crear un programa nuevo hace falta el nombre")
                continue
            codigo_final = codigo or _generar_codigo_programa(empresa_id, nombre, codigos_usados)
            if codigo and _fold(codigo) in programas_codigo:
                errores.append(f"Fila {fila}: ya existe un programa con el código «{codigo}»")
                continue
            tipo = grupo["tipo"] or "interno"
            programa = ProgramaCapacitacion(
                empresa_id=empresa_id,
                codigo=codigo_final,
                nombre=nombre,
                tipo=tipo,
                descripcion=grupo["descripcion"] or None,
                alcance="puesto" if grupo["puestos"] else "general",
                estado="programado",
                activo=True,
            )
            db.session.add(programa)
            db.session.flush()
            programas_codigo[_fold(codigo_final)] = programa
            programas_nombre.setdefault(_fold(nombre), []).append(programa)
            creados += 1
        else:
            programa = hallado
            if not programa.activo:
                programa.activo = True
            if not programa.nombre and nombre:
                programa.nombre = nombre

        cambio = es_nuevo

        for puesto_txt in grupo["puestos"].values():
            puesto = puestos_idx.get(_fold(puesto_txt))
            if not puesto:
                errores.append(
                    f"Fila {fila}: puesto «{puesto_txt}» no encontrado. "
                    "Tiene que existir en el catálogo de puestos."
                )
                continue
            if _agregar_puesto_si_falta(programa, puesto.id):
                puestos_agregados += 1
                cambio = True
                for curso_id in _cursos_del_programa(programa):
                    _ensure_requisito(empresa_id, puesto.id, curso_id)

        for plan_data in grupo["planes"].values():
            plan, creado_plan = _agregar_plan_si_falta(programa, plan_data["nombre"])
            if creado_plan:
                planes_agregados += 1
                cambio = True
            for curso_txt in plan_data["cursos"]:
                curso = cursos_idx.get(_fold(curso_txt))
                if not curso:
                    errores.append(
                        f"Fila {fila}: curso «{curso_txt}» no encontrado. "
                        "Tiene que existir en Cursos y catálogos."
                    )
                    continue
                if _agregar_curso_si_falta(empresa_id, programa, plan, curso):
                    cursos_agregados += 1
                    cambio = True

        if not es_nuevo and cambio:
            actualizados += 1
        elif not es_nuevo and not cambio:
            omitidos += 1

    db.session.commit()
    return {
        "creados": creados,
        "actualizados": actualizados,
        "omitidos": omitidos,
        "planes_agregados": planes_agregados,
        "puestos_agregados": puestos_agregados,
        "cursos_agregados": cursos_agregados,
        "errores": errores,
    }
