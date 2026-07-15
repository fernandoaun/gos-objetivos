"""Exportación de datos de capacitaciones a Excel."""

from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from gos.modulos.capacitacion.services.matriz_analitica_service import (
    MESES_NOMBRES,
    matriz_analitica,
)
from gos.modulos.capacitacion.services.matriz_service import matriz_capacitaciones

# Colores de marca — Green Oil Services (ver gos/static/css/tokens.css)
_BRAND_GREEN = "76B947"
_BRAND_GREEN_DARK = "5F9938"
_BRAND_GREEN_TEXT = "2D5A1E"
_BRAND_YELLOW = "F9E29C"
_BRAND_GREY = "5C5C5C"

_HEADER_FILL = PatternFill(start_color=_BRAND_GREEN_DARK, end_color=_BRAND_GREEN_DARK, fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BAND_FILL = PatternFill(start_color=_BRAND_GREEN, end_color=_BRAND_GREEN, fill_type="solid")

_MARCA_EMPRESA_FONT = Font(color=_BRAND_GREEN_DARK, bold=True, size=12)
_MARCA_TITULO_FONT = Font(color=_BRAND_GREEN_TEXT, bold=True, size=14)
_MARCA_SUBTITULO_FONT = Font(color=_BRAND_GREY, italic=True, size=10)

# Ruta al logo: gos/modulos/capacitacion/services/export_service.py -> gos/static/img/
_LOGO_PATH = Path(__file__).resolve().parents[3] / "static" / "img" / "gos-logo.png"

# Filas reservadas arriba de la tabla para el logo y los títulos.
_MARCA_FILAS = 4
DATOS_INICIO = _MARCA_FILAS + 1

_LEGACY_COLOR_MAP = {
    "verde": "C6EFCE",
    "amarillo": "FFEB9C",
    "rojo": "FFC7CE",
    "azul": "BDD7EE",
    "gris": "D9D9D9",
}


def _aplicar_marca(ws, titulo: str, subtitulo: str = "", ancho_cols: int = 8) -> int:
    """Inserta el logo en el margen superior izquierdo y una banda con los
    colores de la compañía. Devuelve la primera fila disponible para datos."""
    if _LOGO_PATH.is_file():
        try:
            from openpyxl.drawing.image import Image as _XLImage

            img = _XLImage(str(_LOGO_PATH))
            ratio = (img.width / img.height) if img.height else 1.25
            img.height = 54
            img.width = int(round(54 * ratio))
            img.anchor = "A1"
            ws.add_image(img)
            ws.column_dimensions["A"].width = 14
        except Exception:
            pass

    ws.cell(row=1, column=3, value="Green Oil Services").font = _MARCA_EMPRESA_FONT
    ws.cell(row=2, column=3, value=titulo).font = _MARCA_TITULO_FONT
    if subtitulo:
        ws.cell(row=3, column=3, value=subtitulo).font = _MARCA_SUBTITULO_FONT

    for col in range(1, max(ancho_cols, 1) + 1):
        ws.cell(row=_MARCA_FILAS, column=col).fill = _BAND_FILL

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[_MARCA_FILAS].height = 6
    return DATOS_INICIO


_CAL_COLS = [
    ("programados", "Cursos Programados"),
    ("pendientes", "Cursos Pendientes"),
    ("cumplidos", "Cursos Cumplidos"),
    ("pct_cumpl_prog", "% Cumpl./Pr."),
    ("puntuales", "Cumplidos Puntuales"),
    ("pct_punt_prog", "% Punt./Pr."),
    ("pend_vencidos", "Pendientes Vencidos"),
    ("pct_venc_prog", "% Venc./Pr."),
]
_CAL_PCT_COLS = [
    ("pct_pend_sin_vencer", "Pendientes Sin Vencer"),
    ("pct_pend_vencidos", "Pendientes Vencidos"),
    ("pct_cumpl_puntuales", "Cumplidos Puntuales"),
    ("pct_cumpl_no_puntuales", "Cumplidos No Puntuales"),
]

_TABLA_SUB = ("prog", "pdtes", "cumpl", "cumpl_prog")
_TABLA_SUB_LABELS = ("Prog", "Pdtes", "Cumpl", "Cumpl/Prog")


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _fmt_pct(val) -> str | float | int:
    if val is None:
        return ""
    if isinstance(val, float) and 0 <= val <= 1:
        return round(val * 100, 1)
    return val


def exportar_matriz_excel(empresa_id: int, **filtros) -> BytesIO:
    """Exportación legacy (persona × curso)."""
    data = matriz_capacitaciones(empresa_id, **filtros)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz capacitaciones"

    headers = ["Persona", "Legajo", "Sector"] + [c["codigo"] for c in data["columnas"]]
    top = _aplicar_marca(ws, "Matriz de Capacitaciones", ancho_cols=len(headers))
    _write_header(ws, headers, row=top)

    for row_idx, fila in enumerate(data["filas"], top + 1):
        ws.cell(row=row_idx, column=1, value=fila["nombre"])
        ws.cell(row=row_idx, column=2, value=fila.get("legajo") or "")
        ws.cell(row=row_idx, column=3, value=fila.get("sector_nombre") or "")
        for col_idx, col in enumerate(data["columnas"], 4):
            celda = fila["celdas"].get(str(col["id"]), {})
            estado = celda.get("estado", "no_aplica")
            cell = ws.cell(row=row_idx, column=col_idx, value=estado)
            fill_color = _LEGACY_COLOR_MAP.get(celda.get("color", "gris"), "D9D9D9")
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _export_calendario_resumen(ws, data: dict) -> None:
    ws.title = "Capacitaciones"
    anio = data.get("anio", "")
    all_cols = _CAL_COLS + _CAL_PCT_COLS
    top = _aplicar_marca(
        ws, "Matriz Analítica — Calendario", f"Año {anio}", ancho_cols=len(all_cols) + 1
    )

    ws.cell(row=top, column=1, value="Programas")
    ws.cell(row=top + 1, column=1, value="Planes")
    ws.cell(row=top + 2, column=1, value="Personas")

    row_head = top + 3
    ws.cell(row=row_head, column=1, value=anio)
    col = 2
    for _, label in _CAL_COLS:
        ws.cell(row=row_head, column=col, value=label)
        col += 1
    for _, label in _CAL_PCT_COLS:
        ws.cell(row=row_head, column=col, value=label)
        col += 1
    for c in range(1, len(all_cols) + 2):
        cell = ws.cell(row=row_head, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    for i, fila in enumerate(data.get("filas") or [], start=row_head + 1):
        ws.cell(row=i, column=1, value=fila.get("nombre"))
        for j, (key, _) in enumerate(all_cols, start=2):
            ws.cell(row=i, column=j, value=_fmt_pct(fila.get(key)))

    tot = data.get("totales") or {}
    row_tot = row_head + 1 + len(data.get("filas") or [])
    cell_tot = ws.cell(row=row_tot, column=1, value="Total")
    cell_tot.font = Font(bold=True, color=_BRAND_GREEN_TEXT)
    for j, (key, _) in enumerate(all_cols, start=2):
        c = ws.cell(row=row_tot, column=j, value=_fmt_pct(tot.get(key)))
        c.font = Font(bold=True, color=_BRAND_GREEN_TEXT)


def _export_tabla_personas(ws, data: dict) -> None:
    ws.title = "Capacitaciones"
    anio = data.get("anio", "")
    agrupar = data.get("agrupar_por", "persona")
    meses = data.get("meses") or [{"num": i, "nombre": n} for i, n in enumerate(MESES_NOMBRES, start=1)]

    ancho = 1 + (len(meses) + 1) * len(_TABLA_SUB_LABELS)
    top = _aplicar_marca(
        ws,
        "Matriz Analítica — Tabla",
        f"Año {anio} · por {'puesto' if agrupar == 'puesto' else 'persona'}",
        ancho_cols=ancho,
    )

    ws.cell(row=top, column=1, value="Planes:")
    ws.cell(row=top + 1, column=1, value="Puestos" if agrupar == "puesto" else "Personas")

    row_mes = top + 2
    row_sub = top + 3
    ws.cell(row=row_sub, column=1, value=anio)

    col = 2
    for mes in meses:
        c = ws.cell(row=row_mes, column=col, value=mes.get("nombre"))
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _HEADER_ALIGN
        for j, sub in enumerate(_TABLA_SUB_LABELS):
            sc = ws.cell(row=row_sub, column=col + j, value=sub)
            sc.fill = _HEADER_FILL
            sc.font = _HEADER_FONT
            sc.alignment = _HEADER_ALIGN
        col += 4

    ac = ws.cell(row=row_mes, column=col, value="Anual")
    ac.fill = _HEADER_FILL
    ac.font = _HEADER_FONT
    ac.alignment = _HEADER_ALIGN
    for j, sub in enumerate(_TABLA_SUB_LABELS):
        sc = ws.cell(row=row_sub, column=col + j, value=sub)
        sc.fill = _HEADER_FILL
        sc.font = _HEADER_FONT
        sc.alignment = _HEADER_ALIGN

    for row_idx, fila in enumerate(data.get("filas") or [], start=row_sub + 1):
        ws.cell(row=row_idx, column=1, value=fila.get("nombre") or fila.get("persona"))
        meses_data = fila.get("meses") or {}
        col = 2
        for mes in meses:
            m = meses_data.get(str(mes["num"]), {})
            for j, sub in enumerate(_TABLA_SUB):
                ws.cell(row=row_idx, column=col + j, value=_fmt_pct(m.get(sub)))
            col += 4
        anual = meses_data.get("anual", {})
        for j, sub in enumerate(_TABLA_SUB):
            ws.cell(row=row_idx, column=col + j, value=_fmt_pct(anual.get(sub)))


def exportar_matriz_analitica_excel(
    empresa_id: int,
    *,
    vista: str = "tabla",
    anio: int | None = None,
    plan_ids=None,
    tipos=None,
    empresas=None,
    persona_ids=None,
    puesto_ids=None,
    persona_id: int | None = None,
    agrupar_por: str = "persona",
) -> BytesIO:
    """Exporta la matriz analítica según la vista activa."""
    result = matriz_analitica(
        empresa_id,
        vista=vista,
        anio=anio,
        plan_ids=plan_ids,
        tipos=tipos,
        empresas=empresas,
        persona_ids=persona_ids,
        puesto_ids=puesto_ids,
        persona_id=persona_id,
        agrupar_por=agrupar_por,
    )
    data = result["data"]
    wb = openpyxl.Workbook()
    ws = wb.active

    if vista in ("calendar", "calendario"):
        _export_calendario_resumen(ws, data)
    elif vista in ("person", "persona"):
        ws.title = "Persona"
        persona = data.get("persona") or {}
        headers = [
            "Programa",
            "Plan",
            "Curso",
            "Estado",
            "Nota",
            "Horas",
            "Empresa dictada",
            "Vencimiento",
        ]
        top = _aplicar_marca(
            ws,
            "Matriz Analítica — Persona",
            persona.get("nombre") or "",
            ancho_cols=len(headers),
        )
        _write_header(ws, headers, row=top)
        row_idx = top + 1
        for prog in data.get("programas") or []:
            for curso in prog.get("cursos") or []:
                ws.cell(row=row_idx, column=1, value=prog.get("nombre"))
                ws.cell(row=row_idx, column=2, value=prog.get("plan_nombre"))
                ws.cell(row=row_idx, column=3, value=curso.get("nombre"))
                ws.cell(row=row_idx, column=4, value=curso.get("estado"))
                ws.cell(row=row_idx, column=5, value=curso.get("nota"))
                ws.cell(row=row_idx, column=6, value=curso.get("horas"))
                ws.cell(row=row_idx, column=7, value=curso.get("empresa"))
                ws.cell(row=row_idx, column=8, value=curso.get("fecha_vencimiento"))
                row_idx += 1
        if row_idx == top + 1:
            ws.cell(row=top + 1, column=1, value=persona.get("nombre"))
    else:
        _export_tabla_personas(ws, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
