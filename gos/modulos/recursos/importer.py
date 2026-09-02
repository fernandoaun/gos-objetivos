"""Importa la planilla de afectación de unidades (Excel).

Hoja esperada: Unidades (o la primera hoja con cabeceras de servicio).

Estructura:
- Fila 3: destinos (servicios / estructura / estado de parque)
- Filas 4–7: cupos necesarios por tipo UL / TR / ST / OT
- Fila 17: código de equipo por destino (H18, GV35, …)
- Filas 19+: unidades (INTERNO, DOMINIO, SERV) con una «O» en el destino
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from gos.modulos.recursos.models import (
    RecAsignacion,
    RecCambio,
    RecCentro,
    RecCupo,
    RecDestino,
    RecUnidad,
    TIPOS_UNIDAD,
)

ESTRUCTURA_SLUGS = {"MANT", "SMA", "GER", "PCL", "RRHH", "ADM", "COMP"}
ESTADO_SLUGS = {"LIBRE", "REPARACION", "FUERA-DE-SER", "FUERA-DE-SERVICIO"}
SKIP_HEADERS = {"TOTAL", "ESTADO", "SERV", "CO", "INTERNO", "DOMINIO"}

PREFIX_TIPO = {
    "UL": "UL",
    "ULC": "UL",
    "TR": "TR",
    "ST": "ST",
    "STR": "ST",
}


def slug(raw: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "-", (raw or "").upper().strip()).strip("-")


def normalizar_codigo(raw: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", (raw or "").strip().upper())


def tipo_desde_interno(interno: str) -> str:
    letters = ""
    for ch in (interno or "").strip().upper():
        if ch.isalpha():
            letters += ch
        else:
            break
    return PREFIX_TIPO.get(letters, "OT")


def _int_cell(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(float(str(value).replace(",", ".").strip()))
    except ValueError:
        return 0


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_mark(value) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in {"O", "X", "1", "SÍ", "SI"}


def _find_unidades_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "unidades":
            return wb[name]
    return wb[wb.sheetnames[0]]


def _grupo_destino(nombre: str) -> str:
    key = slug(nombre)
    if key in ESTADO_SLUGS:
        return "estado"
    if key in ESTRUCTURA_SLUGS:
        return "estructura"
    return "servicio"


def parse_planilla(path: str | Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = _find_unidades_sheet(wb)

    header_row = 3
    cupo_rows = {"UL": 4, "TR": 5, "ST": 6, "OT": 7}
    equipo_row = 17
    first_unit_row = 19

    destinos: list[dict] = []
    used_codigos: set[str] = set()
    last_col = ws.max_column or 1

    for col in range(8, last_col + 1):
        nombre = _text(ws.cell(header_row, col).value)
        if not nombre or slug(nombre) in SKIP_HEADERS:
            continue
        equipo = _text(ws.cell(equipo_row, col).value)
        codigo = slug(nombre) or get_column_letter(col)
        if codigo in used_codigos:
            extra = slug(equipo) or get_column_letter(col)
            codigo = f"{slug(nombre)}-{extra}"
        if codigo in used_codigos:
            codigo = f"{slug(nombre)}-{get_column_letter(col)}"
        used_codigos.add(codigo)
        cupos = {tipo: _int_cell(ws.cell(row, col).value) for tipo, row in cupo_rows.items()}
        destinos.append(
            {
                "codigo": codigo,
                "nombre": nombre,
                "grupo": _grupo_destino(nombre),
                "equipo": equipo or None,
                "orden": len(destinos),
                "columna_excel": get_column_letter(col),
                "col": col,
                "cupos": cupos,
            }
        )

    unidades: list[dict] = []
    warnings: list[str] = []

    for row in range(first_unit_row, (ws.max_row or first_unit_row) + 1):
        interno = _text(ws.cell(row, 3).value)
        dominio = _text(ws.cell(row, 4).value)
        contratista = _text(ws.cell(row, 7).value)
        if not interno:
            continue
        if interno.upper() in SKIP_HEADERS:
            continue
        codigo = normalizar_codigo(interno)
        if not codigo:
            continue
        marks = []
        for d in destinos:
            if _is_mark(ws.cell(row, d["col"]).value):
                marks.append(d["codigo"])
        if len(marks) > 1:
            warnings.append(
                f"{interno}: más de un destino ({', '.join(marks)}); se toma el primero."
            )
        unidades.append(
            {
                "codigo": codigo,
                "interno": re.sub(r"\s+", " ", interno).strip(),
                "dominio": dominio or None,
                "tipo": tipo_desde_interno(interno),
                "contratista": contratista or None,
                "destino_codigo": marks[0] if marks else None,
            }
        )

    return {
        "destinos": destinos,
        "unidades": unidades,
        "warnings": warnings,
    }


def import_planilla(path: str | Path, session: Session, *, user_id: int | None = None) -> dict:
    parsed = parse_planilla(path)
    destinos_data = parsed["destinos"]
    unidades_data = parsed["unidades"]

    destinos_db = {
        d.codigo: d for d in session.execute(select(RecDestino)).scalars().all()
    }
    seen_destinos: set[str] = set()
    for item in destinos_data:
        dest = destinos_db.get(item["codigo"])
        if dest is None:
            dest = RecDestino(codigo=item["codigo"])
            session.add(dest)
            destinos_db[item["codigo"]] = dest
        dest.nombre = item["nombre"]
        dest.grupo = item["grupo"]
        dest.equipo = item["equipo"]
        dest.orden = item["orden"]
        dest.activo = True
        dest.columna_excel = item["columna_excel"]
        seen_destinos.add(item["codigo"])
    session.flush()

    for codigo, dest in destinos_db.items():
        if codigo not in seen_destinos and dest.columna_excel:
            dest.activo = False
            for centro in list(
                session.execute(select(RecCentro).where(RecCentro.destino_id == dest.id)).scalars().all()
            ):
                centro.destino_id = None

    centros_db = {c.codigo: c for c in session.execute(select(RecCentro)).scalars().all()}
    for item in destinos_data:
        dest = destinos_db[item["codigo"]]
        if dest.grupo != "servicio":
            continue
        equipo = (item.get("equipo") or "").strip()
        if not equipo or equipo.upper() == "GOS":
            for centro in list(
                session.execute(select(RecCentro).where(RecCentro.destino_id == dest.id)).scalars().all()
            ):
                centro.destino_id = None
            dest.equipo = None
            continue
        centro = centros_db.get(equipo)
        if centro is None:
            centro = RecCentro(codigo=equipo, nombre=equipo, activo=True)
            session.add(centro)
            centros_db[equipo] = centro
        else:
            centro.activo = True
        if centro.destino_id not in (None, dest.id):
            previo = session.get(RecDestino, centro.destino_id)
            if previo is not None and (previo.equipo or "").strip() == equipo:
                previo.equipo = None
        session.flush()
        for otro in list(
            session.execute(
                select(RecCentro).where(
                    RecCentro.destino_id == dest.id,
                    RecCentro.id != centro.id,
                )
            ).scalars().all()
        ):
            otro.destino_id = None
        centro.destino_id = dest.id
        dest.equipo = equipo
    session.flush()

    equipos_centro = {
        normalizar_codigo(d.get("equipo") or "")
        for d in destinos_data
        if d.get("equipo") and str(d["equipo"]).upper() != "GOS"
    }
    equipos_centro.discard("")
    unidades_db = {
        u.codigo: u for u in session.execute(select(RecUnidad)).scalars().all()
    }
    seen_unidades: set[str] = set()
    for item in unidades_data:
        unidad = unidades_db.get(item["codigo"])
        if unidad is None:
            unidad = RecUnidad(codigo=item["codigo"])
            session.add(unidad)
            unidades_db[item["codigo"]] = unidad
        unidad.interno = item["interno"]
        unidad.dominio = item["dominio"]
        unidad.tipo = item["tipo"]
        unidad.contratista = item["contratista"]
        unidad.es_centro = bool(unidad.es_centro) or unidad.codigo in equipos_centro
        unidad.activo = True
        unidad.updated_at = datetime.utcnow()
        seen_unidades.add(item["codigo"])
    session.flush()

    ids_excel = {destinos_db[codigo].id for codigo in seen_destinos}
    cupos_existentes = list(session.execute(select(RecCupo)).scalars().all())
    for cupo in cupos_existentes:
        if cupo.destino_id in ids_excel:
            session.delete(cupo)
    session.flush()

    for item in destinos_data:
        dest = destinos_db[item["codigo"]]
        for tipo in TIPOS_UNIDAD:
            necesarias = int(item["cupos"].get(tipo) or 0)
            session.add(RecCupo(destino_id=dest.id, tipo=tipo, necesarias=necesarias))

    asignaciones = {
        a.unidad_id: a for a in session.execute(select(RecAsignacion)).scalars().all()
    }
    assigned = 0
    for item in unidades_data:
        unidad = unidades_db[item["codigo"]]
        actual = asignaciones.get(unidad.id)
        dest_codigo = item["destino_codigo"]
        if not dest_codigo:
            if actual is not None:
                session.delete(actual)
            continue
        dest = destinos_db[dest_codigo]
        if actual is None:
            actual = RecAsignacion(unidad_id=unidad.id, destino_id=dest.id)
            session.add(actual)
            asignaciones[unidad.id] = actual
        else:
            actual.destino_id = dest.id
        actual.updated_at = datetime.utcnow()
        actual.updated_by = user_id
        assigned += 1

    session.add(
        RecCambio(
            user_id=user_id,
            accion="importar",
            entidad="planilla",
            resumen=(
                f"Importó planilla: {len(seen_unidades)} unidades, "
                f"{assigned} asignadas, {len(seen_destinos)} destinos"
            ),
        )
    )
    session.commit()
    return {
        "destinos": len(seen_destinos),
        "unidades": len(seen_unidades),
        "asignadas": assigned,
        "sin_asignar": len(seen_unidades) - assigned,
        "warnings": parsed["warnings"],
    }
