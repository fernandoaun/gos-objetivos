import os
import re
from pathlib import Path

from gos.extensions import db
from gos.modulos.objetivos.models.kpi import KPI_AGREGACION_TIPOS, KpiIndicador

EXCEL_KPI_DEFAULT = Path(
    r"c:\Users\ferna\OneDrive\GOS\Gestion de Indicadores"
    r"\GESTION  INDICADORES GOS 2026 para ir completando.xlsx"
)
HOJA_AVANCE = "Avance KPI 2026"
HOJA_DEFINICION = "KPI "


def _excel_path() -> Path:
    custom = os.environ.get("GOS_KPI_EXCEL_PATH")
    if custom:
        return Path(custom)
    return EXCEL_KPI_DEFAULT


def excel_path() -> str:
    return str(_excel_path())


def _cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_num(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def _objetivo_desde_codigo(codigo: str) -> str | None:
    match = re.match(r"KPI-(\d+)-", codigo or "")
    if not match:
        return None
    return f"OE-{int(match.group(1)):02d}"


def _tipo_agregacion_desde_formula(formula: str | None) -> str:
    if not formula:
        return "promedio"
    if "Σ" in formula:
        return "suma"
    if "%" in formula:
        return "promedio"
    return "ultimo"


def _normalizar_tipo_agregacion(valor: str | None) -> str:
    if valor and valor in KPI_AGREGACION_TIPOS:
        return valor
    return "promedio"


def siguiente_numero_kpi(empresa_id: int) -> int:
    max_num = (
        db.session.query(db.func.max(KpiIndicador.numero))
        .filter_by(empresa_id=empresa_id, activo=True)
        .scalar()
    )
    return (max_num or 0) + 1


def _valores_mes_desde_fila(row, start_col: int = 9) -> dict[str, float]:
    valores: dict[str, float] = {}
    for mes in range(1, 13):
        num = _cell_num(row[start_col + mes - 1].value)
        if num is not None:
            valores[str(mes)] = num
    return valores


def _cargar_formulas_maestro(wb) -> dict[str, str]:
    if HOJA_DEFINICION not in wb.sheetnames:
        return {}
    ws = wb[HOJA_DEFINICION]
    formulas: dict[str, str] = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        codigo = _cell_str(row[2].value)
        if not codigo:
            continue
        formula = _cell_str(row[5].value)
        if formula:
            formulas[codigo] = formula
    return formulas


def _valores_mes_ordenados(kpi: KpiIndicador) -> list[float]:
    valores = kpi.valores_mes or {}
    return [
        float(v)
        for _, v in sorted(valores.items(), key=lambda item: int(item[0]))
        if v is not None and str(v).strip() != ""
    ]


def _calcular_agregado(tipo: str, nums: list[float]) -> float | None:
    if not nums:
        return None
    if tipo == "suma":
        return sum(nums)
    if tipo in ("ultimo", "cumplimiento"):
        return nums[-1]
    if tipo == "promedio":
        return sum(nums) / len(nums)
    return sum(nums) / len(nums)


def calcular_metricas(kpi: KpiIndicador) -> dict:
    nums = _valores_mes_ordenados(kpi)
    tipo = _normalizar_tipo_agregacion(kpi.tipo_agregacion)
    agregado = _calcular_agregado(tipo, nums)

    avance = None
    if agregado is not None and kpi.meta_2026_num not in (None, 0):
        avance = agregado / kpi.meta_2026_num

    if not nums and kpi.meta_2026_num is None:
        estado = "Sin datos"
    elif avance is None:
        estado = "Sin datos"
    elif avance >= 1.0:
        estado = "En meta"
    else:
        estado = "Fuera de meta"

    return {
        "agregado": agregado,
        "avance": avance,
        "estado": estado,
    }


def limpiar_etiquetas_grupo(empresa_id: int) -> None:
    """Quita etiquetas de sección importadas del Excel (ej. GOS CONNECT) de la columna N°."""
    actualizados = (
        KpiIndicador.query.filter_by(empresa_id=empresa_id, activo=True)
        .filter(KpiIndicador.grupo.isnot(None))
        .update({KpiIndicador.grupo: None}, synchronize_session=False)
    )
    if actualizados:
        db.session.commit()


def listar_kpis(empresa_id: int) -> list[KpiIndicador]:
    return (
        KpiIndicador.query.filter_by(empresa_id=empresa_id, activo=True)
        .order_by(KpiIndicador.orden, KpiIndicador.codigo)
        .all()
    )


def listar_kpis_con_metricas(empresa_id: int) -> list[dict]:
    filas = []
    for kpi in listar_kpis(empresa_id):
        metricas = calcular_metricas(kpi)
        filas.append({"kpi": kpi, **metricas})
    return filas


def obtener_kpi(empresa_id: int, kpi_id: int) -> KpiIndicador | None:
    return KpiIndicador.query.filter_by(
        empresa_id=empresa_id, id=kpi_id, activo=True
    ).first()


def actualizar_kpi(
    empresa_id: int,
    kpi_id: int,
    *,
    indicador: str | None = None,
    numero: str | None = None,
    responsable: str | None = None,
    medio: str | None = None,
    resultado_2025: str | None = None,
    meta_2026: str | None = None,
    tipo_agregacion: str | None = None,
    observacion: str | None = None,
    valores_mes: dict | None = None,
) -> KpiIndicador:
    kpi = obtener_kpi(empresa_id, kpi_id)
    if not kpi:
        raise ValueError("KPI no encontrado.")

    if indicador is not None:
        texto = indicador.strip()
        if not texto:
            raise ValueError("El indicador no puede estar vacío.")
        kpi.indicador = texto
    if numero is not None:
        num = _cell_num(numero)
        kpi.numero = int(num) if num is not None else None
    if responsable is not None:
        kpi.responsable = responsable.strip() or None
    if medio is not None:
        kpi.medio = medio.strip() or None
    if resultado_2025 is not None:
        kpi.resultado_2025 = resultado_2025.strip() or None
    if meta_2026 is not None:
        kpi.meta_2026 = meta_2026.strip() or None
        kpi.meta_2026_num = _cell_num(kpi.meta_2026)
    if tipo_agregacion is not None:
        kpi.tipo_agregacion = _normalizar_tipo_agregacion(tipo_agregacion)
    if observacion is not None:
        kpi.observacion = observacion.strip() or None
    if valores_mes is not None:
        limpio: dict[str, float] = {}
        for mes in range(1, 13):
            raw = valores_mes.get(str(mes), valores_mes.get(mes))
            if raw is None or str(raw).strip() == "":
                continue
            num = _cell_num(raw)
            if num is not None:
                limpio[str(mes)] = num
        kpi.valores_mes = limpio

    db.session.commit()
    return kpi


def crear_kpi(
    empresa_id: int,
    *,
    codigo: str,
    indicador: str,
    numero: str | None = None,
    responsable: str | None = None,
    medio: str | None = None,
    resultado_2025: str | None = None,
    meta_2026: str | None = None,
    tipo_agregacion: str | None = None,
    observacion: str | None = None,
    valores_mes: dict | None = None,
) -> KpiIndicador:
    codigo_limpio = codigo.strip().upper()
    indicador_limpio = indicador.strip()
    if not codigo_limpio:
        raise ValueError("El código KPI es obligatorio.")
    if not indicador_limpio:
        raise ValueError("El indicador es obligatorio.")

    existente = KpiIndicador.query.filter_by(
        empresa_id=empresa_id, codigo=codigo_limpio, activo=True
    ).first()
    if existente:
        raise ValueError(f"Ya existe un KPI con código {codigo_limpio}.")

    max_orden = (
        db.session.query(db.func.max(KpiIndicador.orden))
        .filter_by(empresa_id=empresa_id)
        .scalar()
        or 0
    )
    meta_texto = meta_2026.strip() if meta_2026 else None
    num_orden = _cell_num(numero)
    if num_orden is None:
        num_orden = float(siguiente_numero_kpi(empresa_id))

    kpi = KpiIndicador(
        empresa_id=empresa_id,
        numero=int(num_orden),
        codigo=codigo_limpio,
        indicador=indicador_limpio,
        objetivo_codigo=_objetivo_desde_codigo(codigo_limpio),
        responsable=responsable.strip() if responsable else None,
        medio=medio.strip() if medio else None,
        resultado_2025=resultado_2025.strip() if resultado_2025 else None,
        meta_2026=meta_texto,
        meta_2026_num=_cell_num(meta_texto) if meta_texto else None,
        valores_mes={},
        tipo_agregacion=_normalizar_tipo_agregacion(tipo_agregacion),
        observacion=observacion.strip() if observacion else None,
        orden=max_orden + 1,
        activo=True,
    )

    if valores_mes:
        limpio: dict[str, float] = {}
        for mes in range(1, 13):
            raw = valores_mes.get(str(mes), valores_mes.get(mes))
            if raw is None or str(raw).strip() == "":
                continue
            num = _cell_num(raw)
            if num is not None:
                limpio[str(mes)] = num
        kpi.valores_mes = limpio

    db.session.add(kpi)
    db.session.commit()
    return kpi


def eliminar_kpi(empresa_id: int, kpi_id: int) -> str:
    kpi = obtener_kpi(empresa_id, kpi_id)
    if not kpi:
        raise ValueError("KPI no encontrado.")
    codigo = kpi.codigo
    kpi.activo = False
    db.session.commit()
    return codigo


def importar_desde_excel(empresa_id: int, *, reemplazar: bool = False) -> int:
    path = _excel_path()
    if not path.is_file():
        raise ValueError(f"No se encontró el archivo Excel: {path}")

    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("Falta instalar openpyxl para importar KPI desde Excel.") from exc

    existentes = KpiIndicador.query.filter_by(empresa_id=empresa_id, activo=True).count()
    if existentes > 0 and not reemplazar:
        raise ValueError(
            "Ya hay KPI cargados. Reimportá con confirmación para reemplazar los datos."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    if HOJA_AVANCE not in wb.sheetnames:
        raise ValueError(f'No se encontró la hoja "{HOJA_AVANCE}" en el Excel.')

    formulas = _cargar_formulas_maestro(wb)
    ws = wb[HOJA_AVANCE]

    if reemplazar:
        KpiIndicador.query.filter_by(empresa_id=empresa_id).delete()
        db.session.flush()

    orden = 0
    importados = 0
    siguiente_num = 1

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        codigo = _cell_str(row[2].value)
        indicador = _cell_str(row[3].value)
        if not codigo or not indicador:
            continue

        numero = _cell_num(row[1].value)
        if numero is None:
            numero = float(siguiente_num)
            siguiente_num += 1
        else:
            siguiente_num = int(numero) + 1

        meta_texto = _cell_str(row[7].value)
        formula = formulas.get(codigo)
        orden += 1

        db.session.add(
            KpiIndicador(
                empresa_id=empresa_id,
                numero=int(numero),
                codigo=codigo,
                indicador=indicador,
                objetivo_codigo=_objetivo_desde_codigo(codigo),
                responsable=_cell_str(row[4].value),
                medio=_cell_str(row[5].value),
                resultado_2025=_cell_str(row[6].value),
                meta_2026=meta_texto,
                meta_2026_num=_cell_num(meta_texto),
                valores_mes=_valores_mes_desde_fila(row),
                tipo_agregacion=_tipo_agregacion_desde_formula(formula),
                observacion=_cell_str(row[24].value),
                grupo=None,
                orden=orden,
                activo=True,
            )
        )
        importados += 1

    if importados == 0:
        raise ValueError("El Excel no contiene filas de KPI en la hoja de avance.")

    db.session.commit()
    return importados
