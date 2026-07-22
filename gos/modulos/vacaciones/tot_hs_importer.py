"""Importación Tot Hs. desde Excel de resumen por período.

Formato esperado (como «archivo (14).xlsx»):
  - Fila de título con rango: «21/12/2025 al 20/07/2026»
  - Encabezados: Nombre, Servicio, Centro, Cliente, Tipo Servicio, Total Horas, …
  - Una fila por (nombre, servicio, centro) con totales del período.

Períodos nuevos se agregan; si el mismo rango ya existe, se reemplaza completo.
"""

from __future__ import annotations

import math
import re
from datetime import date, datetime

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from gos.modulos.vacaciones.models import TotHs

_PERIOD_RE = re.compile(
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*al\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    re.IGNORECASE,
)

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "empleado": ("nombre", "empleado", "persona", "trabajador"),
    "servicio": ("servicio",),
    "centro": ("centro",),
    "cliente": ("cliente",),
    "tipo_servicio": ("tipo servicio", "tipo_servicio", "tipo"),
    "total_horas": ("total horas", "total_horas", "horas", "tot hs", "total hs"),
    "hs_viaje": ("hs.viaje", "hs viaje", "hs_viaje", "horas viaje"),
    "hs50": ("hs.50", "hs 50", "hs50", "hs 50%", "50%"),
    "hs_noc": ("hs.noct", "hs noct", "hs_noc", "hs nocturnas", "nocturnas"),
    "hs_noc50": ("hs.noct50", "hs noct50", "hs_noc50", "hs noc 50", "noct50"),
    "hs100": ("hs.100", "hs 100", "hs100", "hs 100%", "100%"),
    "viandas": ("viandas", "vianda"),
    "v_desayuno": ("desayunos", "desayuno", "v_desayuno", "v desayuno"),
    "d_normales": ("d.normales", "d normales", "d_normales", "dias normales"),
    "ausente": ("ausente", "ausencias"),
    "fr_trabajados": ("fr.trabajados", "fr trabajados", "fr_trabajados"),
    "feriados": ("feriado", "feriados"),
    "enfermedad": ("enfermo", "enfermedad"),
    "traslado": ("traslado", "traslados"),
    "vacaciones": ("vacaciones", "vacacion"),
    "licencia": ("licencia", "licencias"),
    "accidente": ("accidente", "accidentes"),
    "francos_comp": ("fr.compens", "fr compens", "francos_comp", "francos comp"),
    "total_hs_viaje": (
        "totalhs+hsviaje",
        "total hs+hs viaje",
        "totalhs + hsviaje",
        "total horas + viaje",
    ),
}

_FLOAT_FIELDS = frozenset(
    {
        "total_horas", "hs_viaje", "hs50", "hs_noc", "hs_noc50", "hs100",
        "viandas", "v_desayuno", "d_normales", "ausente", "fr_trabajados",
        "feriados", "enfermedad", "traslado", "vacaciones", "licencia",
        "accidente", "francos_comp", "total_hs_viaje",
    }
)


def _norm(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    for old, new in (
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n"),
    ):
        text = text.replace(old, new)
    text = text.replace("_", " ").replace(".", " ")
    return " ".join(text.split())


def _to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return 0.0 if (isinstance(value, float) and math.isnan(value)) else float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_period(text) -> tuple[date, date] | None:
    if text is None:
        return None
    if isinstance(text, (date, datetime)):
        return None
    raw = str(text).strip()
    m = _PERIOD_RE.search(raw)
    if not m:
        return None
    d1, d2 = _to_date(m.group(1)), _to_date(m.group(2))
    if d1 is None or d2 is None:
        return None
    if d2 < d1:
        d1, d2 = d2, d1
    return d1, d2


def _match_field(header) -> str | None:
    h = _norm(header)
    if not h:
        return None
    # Evitar mapear "TotalHs+HsViaje" a total_horas.
    if "viaje" in h and ("total" in h or "+" in str(header)):
        return "total_hs_viaje"
    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if h == _norm(alias):
                return field
    if "total" in h and "hora" in h and "viaje" not in h:
        return "total_horas"
    return None


def _map_headers(row: tuple) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row or ()):
        field = _match_field(cell)
        if field and field not in mapping:
            mapping[field] = idx
    if "empleado" not in mapping or "total_horas" not in mapping:
        return None
    return mapping


def _detect_layout(ws) -> tuple[date, date, int, dict[str, int]] | None:
    """Devuelve (desde, hasta, header_row_0based, mapping)."""
    rows = list(ws.iter_rows(max_row=12, values_only=True))
    period: tuple[date, date] | None = None
    for row in rows[:6]:
        for cell in row or ():
            period = _parse_period(cell)
            if period:
                break
        if period:
            break
    if not period:
        return None

    for i, row in enumerate(rows):
        mapping = _map_headers(row or ())
        if mapping:
            return period[0], period[1], i, mapping
    return None


def _score_sheet(ws) -> int:
    layout = _detect_layout(ws)
    if not layout:
        return 0
    _d, _h, _i, mapping = layout
    return 10 + len(mapping)


def _pick_sheet(wb) -> str | None:
    best_name = None
    best_score = 0
    for name in wb.sheetnames:
        score = _score_sheet(wb[name])
        # Preferir hojas llamadas Total / Tot Hs solo como desempate.
        if _norm(name) in ("total", "tot hs", "tot hs.", "horas"):
            score += 3
        if score > best_score:
            best_score = score
            best_name = name
    return best_name if best_score >= 10 else None


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


def import_tot_hs_excel(filepath: str, db: Session) -> dict:
    """Importa resumen por período. Pisá el período si ya existía."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ValueError(
            "No se pudo leer el archivo. Asegurate de que sea un Excel válido "
            "(.xlsx o .xlsm)."
        ) from exc

    try:
        sheet_name = _pick_sheet(wb)
        if not sheet_name:
            raise ValueError(
                "No se encontró una planilla de Tot Hs. "
                "Tiene que tener un título con el rango de fechas "
                "(ej. «21/12/2025 al 20/07/2026») y columnas Nombre / Total Horas."
            )
        ws = wb[sheet_name]
        layout = _detect_layout(ws)
        if layout is None:
            raise ValueError(
                "La hoja tiene datos pero no se pudo leer el período o los encabezados."
            )
        desde, hasta, header_idx, mapping = layout

        records: list[dict] = []
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            empleado = _to_str(_cell(row, mapping["empleado"]))
            if not empleado:
                continue
            servicio = _to_str(_cell(row, mapping["servicio"])) if "servicio" in mapping else ""
            centro = _to_str(_cell(row, mapping["centro"])) if "centro" in mapping else ""
            rec = {
                "periodo_desde": desde,
                "periodo_hasta": hasta,
                "empleado": empleado,
                "servicio": servicio,
                "centro": centro,
            }
            for field in _FLOAT_FIELDS | {"cliente", "tipo_servicio"}:
                if field in ("empleado", "servicio", "centro"):
                    continue
                idx = mapping.get(field)
                if idx is None:
                    if field in _FLOAT_FIELDS:
                        rec[field] = 0.0
                    else:
                        rec[field] = None
                    continue
                val = _cell(row, idx)
                if field in _FLOAT_FIELDS:
                    rec[field] = _to_float(val)
                else:
                    rec[field] = _to_str(val) or None
            records.append(rec)
    finally:
        wb.close()

    if not records:
        raise ValueError(
            f"Se detectó el período {desde.isoformat()} → {hasta.isoformat()} "
            "pero no había filas con nombre de empleado."
        )

    existing_count = db.execute(
        select(TotHs.id).where(
            TotHs.periodo_desde == desde,
            TotHs.periodo_hasta == hasta,
        ).limit(1)
    ).first()
    was_update = existing_count is not None

    db.execute(
        delete(TotHs).where(
            TotHs.periodo_desde == desde,
            TotHs.periodo_hasta == hasta,
        )
    )
    db.bulk_insert_mappings(TotHs, records)
    db.commit()

    return {
        "registros": len(records),
        "registros_nuevos": 0 if was_update else len(records),
        "registros_actualizados": len(records) if was_update else 0,
        "periodo_reemplazado": was_update,
        "fecha_min": desde.isoformat(),
        "fecha_max": hasta.isoformat(),
        "periodo_label": f"{desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}",
        "personas": len({r["empleado"] for r in records}),
        "vacaciones": 0,
        "vacaciones_nuevas": 0,
        "vacaciones_actualizadas": 0,
        "errores": [],
    }
