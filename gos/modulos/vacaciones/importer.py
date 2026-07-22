"""Importación de vacaciones desde Excel.

La lectura se hace por *streaming* (openpyxl en modo read-only) para que el
uso de memoria sea plano sin importar el tamaño del archivo: las filas se
procesan y se guardan de a lotes, sin materializar nunca la hoja completa.
Esto evita que un archivo grande agote la memoria del servidor (OOM).
"""

from __future__ import annotations

import ast
import math
import re
from datetime import date, datetime

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from gos.modulos.vacaciones.models import Registro, Vacacion

COLS_TOTAL = [
    "fecha", "empleado", "sector", "servicio", "centro", "situacion",
    "total_horas", "hs_viaje", "hs50", "hs_noc", "hs_noc50", "hs100",
    "viandas", "v_desayuno", "d_normales", "ausente", "fr_trabajados",
    "feriados", "enfermedad", "traslado", "vacaciones", "licencia",
    "suspension", "accidente", "francos_comp",
]

# Columnas de texto y numéricas de la hoja TOTAL (según el modelo Registro).
_TEXT_COLS = frozenset(COLS_TOTAL[2:6])          # sector, servicio, centro, situacion
_FLOAT_COLS = frozenset(COLS_TOTAL[6:12])         # total_horas … hs100
_INT_COLS = frozenset(COLS_TOTAL[12:])            # viandas … francos_comp

CHUNK_SIZE = 500

_SET_REGISTRO = {c: c for c in COLS_TOTAL[2:]}    # columnas actualizables en el upsert


# ── Coerción de valores ────────────────────────────────────────────────────


def _to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return 0.0 if (isinstance(value, float) and math.isnan(value)) else float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return 0.0


def _to_int(value) -> int:
    try:
        return int(_to_float(value))
    except (ValueError, OverflowError):
        return 0


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def _eval_simple_formula(value) -> float | None:
    """Evalúa sumas literales (=SUM(7,3) / =7+3) cuando no hay valor en caché."""
    if not _is_formula(value):
        return None
    expr = value.strip()[1:].strip()
    if not expr:
        return None
    sum_match = re.fullmatch(
        r"(?i)(?:SUM|SUMA)\s*\((.+)\)",
        expr,
    )
    if sum_match:
        parts = re.split(r"[,;]", sum_match.group(1))
        total = 0.0
        for part in parts:
            token = part.strip()
            if not token:
                continue
            try:
                total += float(token)
            except ValueError:
                return None
        return total
    # Solo aritmética con literales numéricos (sin referencias a celdas).
    if not re.fullmatch(r"[\d\s+\-*/().]+", expr):
        return None
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError:
        return None

    def _check(node: ast.AST) -> bool:
        if isinstance(node, ast.Expression):
            return _check(node.body)
        if isinstance(node, ast.Constant):
            return isinstance(node.value, (int, float)) and not isinstance(node.value, bool)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            return _check(node.operand)
        if isinstance(node, ast.BinOp) and isinstance(
            node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)
        ):
            return _check(node.left) and _check(node.right)
        return False

    if not _check(tree):
        return None
    try:
        return float(eval(compile(tree, "<excel>", "eval"), {"__builtins__": {}}, {}))
    except Exception:
        return None


def _resolve_excel_number(raw, cached=None):
    """Número de celda: valor directo, caché de fórmula (data_only) o suma literal."""
    if isinstance(raw, bool):
        return float(raw)
    if isinstance(raw, (int, float)):
        return 0.0 if (isinstance(raw, float) and math.isnan(raw)) else float(raw)
    if _is_formula(raw):
        if cached is not None and cached != "" and not _is_formula(cached):
            return cached
        evaluated = _eval_simple_formula(raw)
        if evaluated is not None:
            return evaluated
        return None
    if raw in (None, ""):
        if cached is not None and cached != "" and not _is_formula(cached):
            return cached
        return None
    return raw


def _to_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


# ── Upsert (Postgres / SQLite) ─────────────────────────────────────────────


def _insert(model, session: Session):
    if session.get_bind().dialect.name == "postgresql":
        from sqlalchemy.dialects.postgresql import insert
    else:
        from sqlalchemy.dialects.sqlite import insert
    return insert(model)


def _open_workbook(filepath: str):
    """Abre el .xlsx/.xlsm en modo read-only (lectura por streaming)."""
    try:
        return openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:  # archivo corrupto, .xls disfrazado, etc.
        raise ValueError(
            "No se pudo leer el archivo. Asegurate de que sea un Excel válido "
            "(.xlsx o .xlsm)."
        ) from exc


# ── Hoja de horas diarias (streaming) ───────────────────────────────────────

# Alias de encabezados → campo del modelo (lo que importa es el contenido).
_TOTAL_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "fecha": ("fecha", "date", "dia", "día"),
    "empleado": ("empleado", "nombre", "persona", "trabajador", "legajo nombre"),
    "sector": ("sector", "area", "área", "departamento"),
    "servicio": ("servicio",),
    "centro": ("centro", "centro de costo", "c.costo", "ccosto"),
    "situacion": ("situacion", "situación", "estado"),
    "total_horas": (
        "total_horas", "total horas", "horas", "hs totales", "total hs",
        "tot hs", "hs total", "horas totales",
    ),
    "hs_viaje": ("hs_viaje", "hs viaje", "horas viaje", "viaje"),
    "hs50": ("hs50", "hs 50", "hs 50%", "horas 50", "50%"),
    "hs_noc": ("hs_noc", "hs noc", "hs nocturnas", "nocturnas", "hs noche"),
    "hs_noc50": ("hs_noc50", "hs noc50", "hs noc 50", "noc 50"),
    "hs100": ("hs100", "hs 100", "hs 100%", "horas 100", "100%"),
    "viandas": ("viandas", "vianda"),
    "v_desayuno": ("v_desayuno", "v desayuno", "desayuno"),
    "d_normales": ("d_normales", "d normales", "dias normales", "días normales"),
    "ausente": ("ausente", "ausencias", "ausencia"),
    "fr_trabajados": ("fr_trabajados", "fr trabajados", "francos trabajados"),
    "feriados": ("feriados", "feriado"),
    "enfermedad": ("enfermedad", "enfermo"),
    "traslado": ("traslado", "traslados"),
    "vacaciones": ("vacaciones", "vacacion", "vacación"),
    "licencia": ("licencia", "licencias"),
    "suspension": ("suspension", "suspensión"),
    "accidente": ("accidente", "accidentes"),
    "francos_comp": ("francos_comp", "francos comp", "francos compensatorios"),
}


def _match_total_field(header: str) -> str | None:
    """Devuelve el campo del modelo si el encabezado coincide con algún alias."""
    h = _norm_header(header)
    if not h:
        return None
    # Evitar "fecha ingreso" de la planilla de vacaciones.
    if h.startswith("fecha ingreso") or h.startswith("fecha de ingreso"):
        return None
    for field, aliases in _TOTAL_HEADER_ALIASES.items():
        for alias in aliases:
            if h == alias or h.replace(" ", "_") == alias.replace(" ", "_"):
                return field
    if h in ("fecha", "date", "dia", "día"):
        return "fecha"
    if h in ("empleado", "nombre", "persona", "trabajador"):
        return "empleado"
    if "total" in h and "hora" in h:
        return "total_horas"
    if h in ("horas", "hs", "total hs", "tot hs"):
        return "total_horas"
    return None


def _map_total_headers(header_row: tuple) -> dict[str, int] | None:
    """Mapea campo → índice de columna. Requiere fecha + empleado + señal de horas."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row or ()):
        field = _match_total_field(cell)
        if field and field not in mapping:
            mapping[field] = idx
    if "fecha" not in mapping or "empleado" not in mapping:
        return None
    # Distinguir de planilla de vacaciones (empleado/sector sin horas diarias).
    hours_markers = {
        "total_horas", "hs_viaje", "hs50", "hs_noc", "hs_noc50", "hs100",
        "d_normales", "ausente", "viandas",
    }
    if not (hours_markers & mapping.keys()):
        return None
    return mapping


def _score_total_header(header_row: tuple) -> int:
    mapping = _map_total_headers(header_row)
    if not mapping:
        return 0
    score = 10  # fecha + empleado + marcador de horas
    for field in ("total_horas", "sector", "hs50", "hs100", "ausente", "vacaciones"):
        if field in mapping:
            score += 2
    score += min(len(mapping), 12)
    return score


def _find_total_layout(ws) -> tuple[int, dict[str, int]] | None:
    """Busca la fila de encabezado (primeras 8) y el mapeo de columnas."""
    best: tuple[int, dict[str, int], int] | None = None
    for i, row in enumerate(ws.iter_rows(max_row=8, values_only=True)):
        score = _score_total_header(row or ())
        if score < 10:
            continue
        mapping = _map_total_headers(row or ())
        if mapping is None:
            continue
        if best is None or score > best[2]:
            best = (i, mapping, score)
    if best is None:
        return None
    return best[0], best[1]


def _detect_total_sheet(wb) -> str | None:
    """Elige la hoja de horas diarias por contenido (no por el nombre)."""
    preferred = {p.upper() for p in ("TOTAL", "TOT HS", "TOT HS.", "HORAS", "REGISTROS")}
    candidates: list[tuple[int, str]] = []

    for name in wb.sheetnames:
        best_score = 0
        for row in wb[name].iter_rows(max_row=8, values_only=True):
            score = _score_total_header(row or ())
            if score > best_score:
                best_score = score
        if best_score < 10:
            continue
        if name.strip().upper() in preferred or name.strip().upper().startswith("TOTAL"):
            best_score += 5
        candidates.append((best_score, name))

    if not candidates:
        return None
    candidates.sort(key=lambda x: (-x[0], x[1]))
    return candidates[0][1]


def _resolve_total_sheet(filepath: str) -> tuple[str, int, dict[str, int]] | None:
    """Abre el libro y resuelve (nombre_hoja, fila_header_0based, mapping)."""
    wb = _open_workbook(filepath)
    try:
        sheet_name = _detect_total_sheet(wb)
        if not sheet_name:
            return None
        layout = _find_total_layout(wb[sheet_name])
        if layout is None:
            return None
        header_idx, mapping = layout
        return sheet_name, header_idx, mapping
    finally:
        wb.close()


def _iter_total_keys(filepath: str):
    """Primera pasada liviana: solo (fecha, empleado) para contar y acotar rango."""
    resolved = _resolve_total_sheet(filepath)
    if resolved is None:
        return
    sheet_name, header_idx, mapping = resolved
    fecha_i = mapping["fecha"]
    emp_i = mapping["empleado"]
    max_col = max(fecha_i, emp_i) + 1
    wb = _open_workbook(filepath)
    try:
        rows = wb[sheet_name].iter_rows(min_col=1, max_col=max_col, values_only=True)
        for _ in range(header_idx + 1):
            next(rows, None)
        for row in rows:
            fecha = _to_date(_cell(row, fecha_i))
            empleado = _to_str(_cell(row, emp_i))
            if fecha is None or not empleado:
                continue
            yield fecha, empleado
    finally:
        wb.close()


def _iter_total_records(filepath: str):
    """Segunda pasada: filas limpias listas para el upsert."""
    resolved = _resolve_total_sheet(filepath)
    if resolved is None:
        return
    sheet_name, header_idx, mapping = resolved
    max_col = max(mapping.values()) + 1
    wb = _open_workbook(filepath)
    try:
        rows = wb[sheet_name].iter_rows(min_col=1, max_col=max_col, values_only=True)
        for _ in range(header_idx + 1):
            next(rows, None)
        for row in rows:
            fecha = _to_date(_cell(row, mapping["fecha"]))
            if fecha is None:
                continue
            empleado = _to_str(_cell(row, mapping["empleado"]))
            if not empleado:
                continue
            rec = {"fecha": fecha, "empleado": empleado}
            for name in COLS_TOTAL[2:]:
                idx = mapping.get(name)
                if idx is None:
                    if name in _INT_COLS:
                        rec[name] = 0
                    elif name in _FLOAT_COLS:
                        rec[name] = 0.0
                    else:
                        rec[name] = None
                    continue
                val = _cell(row, idx)
                if name in _INT_COLS:
                    rec[name] = _to_int(val)
                elif name in _FLOAT_COLS:
                    rec[name] = _to_float(val)
                else:
                    rec[name] = _to_str(val)
            yield rec
    finally:
        wb.close()


def _flush_registros(chunk: list[dict], db: Session) -> None:
    stmt = _insert(Registro, db).values(chunk)
    stmt = stmt.on_conflict_do_update(
        index_elements=["fecha", "empleado"],
        set_={c: getattr(stmt.excluded, c) for c in _SET_REGISTRO},
    )
    db.execute(stmt)


def _import_total(filepath: str, db: Session, result: dict) -> None:
    # Pasada 1: contar nuevos vs. actualizados sin cargar los datos completos.
    keys_en_excel: set[tuple] = set()
    min_date: date | None = None
    max_date: date | None = None
    for fecha, empleado in _iter_total_keys(filepath):
        keys_en_excel.add((fecha, empleado))
        if min_date is None or fecha < min_date:
            min_date = fecha
        if max_date is None or fecha > max_date:
            max_date = fecha

    if min_date is not None:
        result["fecha_min"] = min_date.isoformat()
    if max_date is not None:
        result["fecha_max"] = max_date.isoformat()

    if keys_en_excel:
        existing = db.execute(
            select(Registro.fecha, Registro.empleado).where(
                Registro.fecha >= min_date,
                Registro.fecha <= max_date,
            )
        ).fetchall()
        existing_keys = {(row.fecha, row.empleado) for row in existing}
        result["registros_nuevos"] = len(keys_en_excel - existing_keys)
        result["registros_actualizados"] = len(keys_en_excel & existing_keys)

    # Pasada 2: guardar por lotes (commit por lote para no acumular memoria).
    total = 0
    chunk: list[dict] = []
    for rec in _iter_total_records(filepath):
        chunk.append(rec)
        if len(chunk) >= CHUNK_SIZE:
            _flush_registros(chunk, db)
            db.commit()
            total += len(chunk)
            chunk = []
    if chunk:
        _flush_registros(chunk, db)
        db.commit()
        total += len(chunk)
    result["registros"] = total


def _workbook_has_hours_sheet(filepath: str) -> bool:
    return _resolve_total_sheet(filepath) is not None


# ── Planilla de vacaciones (una fila por empleado, bloques por año) ─────────


def _norm_header(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip().lower()
    for old, new in (
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n"),
    ):
        text = text.replace(old, new)
    return " ".join(text.split())


def _find_year_in_header(value) -> int | None:
    text = _norm_header(value)
    if "vacacion" not in text:
        return None
    for token in text.replace("-", " ").split():
        if token.isdigit() and len(token) == 4:
            year = int(token)
            if 2000 <= year <= 2100:
                return year
    return None


def _detect_planilla_sheet(wb) -> str | None:
    """Elige la hoja de planilla: nombre conocido o detección por encabezados."""
    preferred = ("PLANILLA VACACIONES", "PLANILLA", "VACACIONES")
    for name in preferred:
        if name in wb.sheetnames:
            return name

    best_name = None
    best_score = 0
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(max_row=5, values_only=True))
        flat = [_norm_header(c) for row in rows for c in (row or ())]
        joined = " ".join(flat)
        score = 0
        if any("legajo" in h for h in flat):
            score += 2
        if any(h == "empleado" or h.endswith(" empleado") for h in flat):
            score += 2
        if "vacacion" in joined:
            score += 2
        if any(_find_year_in_header(c) for row in rows for c in (row or ())):
            score += 2
        if score > best_score:
            best_score = score
            best_name = name
    return best_name if best_score >= 4 else None


def _detect_year_blocks(header_rows: list[tuple]) -> list[tuple[int, int]]:
    """Devuelve [(anio, col_offset_dias_disponibles), ...] desde los encabezados."""
    blocks: list[tuple[int, int]] = []
    seen: set[int] = set()
    for row in header_rows:
        for idx, cell in enumerate(row or ()):
            anio = _find_year_in_header(cell)
            if anio is None or anio in seen:
                continue
            # El título del bloque suele estar sobre "Días disponibles".
            blocks.append((anio, idx))
            seen.add(anio)
    if blocks:
        return sorted(blocks, key=lambda x: x[0])
    # Fallback del layout histórico: disponibles en cols 5 / 9 / 13.
    return [(2023, 5), (2024, 9), (2025, 13)]


def _find_data_start(all_rows: list[tuple]) -> int:
    """Primera fila con legajo numérico + nombre de empleado."""
    for i, row in enumerate(all_rows):
        legajo = _cell(row, 0)
        empleado = _to_str(_cell(row, 1))
        if not empleado:
            continue
        if isinstance(legajo, (int, float)) and not isinstance(legajo, bool):
            return i
        if isinstance(legajo, str) and legajo.strip().isdigit():
            return i
    return min(4, len(all_rows))


def _year_block_has_data(row: tuple, col_offset: int) -> bool:
    return any(_cell(row, col_offset + i) not in (None, "") for i in range(3))


def _clean_comment_text(raw: str | None) -> str | None:
    """Quita el prefijo 'Autor:' típico de comentarios Excel."""
    if not raw:
        return None
    text = str(raw).strip()
    if not text:
        return None
    lines = text.splitlines()
    if len(lines) > 1 and lines[0].strip().endswith(":"):
        text = "\n".join(lines[1:]).strip()
    return text or None


def _cell_comment(cell) -> str | None:
    if cell is None or not getattr(cell, "comment", None):
        return None
    return _clean_comment_text(cell.comment.text)


def _merge_nota(*parts: str | None) -> str | None:
    chunks = [p.strip() for p in parts if p and str(p).strip()]
    return "\n".join(chunks) if chunks else None


def _parse_year_block(
    data_rows: list[tuple], col_offset: int, anio: int
) -> list[dict]:
    """Parseo liviano solo por valores (tests / fallback sin hoja viva)."""
    rows = []
    for row in data_rows:
        empleado = _to_str(_cell(row, 1))
        if not empleado:
            continue
        if not _year_block_has_data(row, col_offset):
            continue
        legajo_raw = _cell(row, 0)
        rows.append(
            {
                "legajo": _to_int(legajo_raw) if legajo_raw not in (None, "") else None,
                "empleado": empleado,
                "fecha_ingreso": _to_date(_cell(row, 2)),
                "sector": _to_str(_cell(row, 3)),
                "anio": anio,
                "dias_disponibles": _to_int(_resolve_excel_number(_cell(row, col_offset))),
                "dias_tomados": _to_int(_resolve_excel_number(_cell(row, col_offset + 1))),
                "dias_pendientes": _to_int(_resolve_excel_number(_cell(row, col_offset + 2))),
                "comentario": None,
                "nota_q": _to_str(_cell(row, 16)),
                "nota_r": _to_str(_cell(row, 17)),
            }
        )
    return rows


def _parse_year_block_ws(ws, data_start: int, col_offset: int, anio: int, ws_data=None) -> list[dict]:
    """Parsea filas de datos con comentarios de celda (Días tomados + Q/R)."""

    def _num(row_idx: int, col_idx: int):
        raw = ws.cell(row_idx, col_idx).value
        cached = ws_data.cell(row_idx, col_idx).value if ws_data is not None else None
        return _to_int(_resolve_excel_number(raw, cached))

    rows = []
    for row_idx in range(data_start + 1, ws.max_row + 1):
        empleado = _to_str(ws.cell(row_idx, 2).value)
        if not empleado:
            continue
        disp_raw = ws.cell(row_idx, col_offset + 1).value
        tom_raw = ws.cell(row_idx, col_offset + 2).value
        pend_raw = ws.cell(row_idx, col_offset + 3).value
        if all(v in (None, "") for v in (disp_raw, tom_raw, pend_raw)):
            continue
        legajo_raw = ws.cell(row_idx, 1).value
        tomados_cell = ws.cell(row_idx, col_offset + 2)
        q_cell = ws.cell(row_idx, 17)
        r_cell = ws.cell(row_idx, 18)
        rows.append(
            {
                "legajo": _to_int(legajo_raw) if legajo_raw not in (None, "") else None,
                "empleado": empleado,
                "fecha_ingreso": _to_date(ws.cell(row_idx, 3).value),
                "sector": _to_str(ws.cell(row_idx, 4).value),
                "anio": anio,
                "dias_disponibles": _num(row_idx, col_offset + 1),
                "dias_tomados": _num(row_idx, col_offset + 2),
                "dias_pendientes": _num(row_idx, col_offset + 3),
                "comentario": _cell_comment(tomados_cell),
                "nota_q": _merge_nota(_to_str(q_cell.value), _cell_comment(q_cell)),
                "nota_r": _merge_nota(_to_str(r_cell.value), _cell_comment(r_cell)),
            }
        )
    return rows


def _import_planilla(filepath: str, db: Session, result: dict) -> None:
    # data_only=False para conservar comentarios (triángulo rojo).
    # data_only=True para leer el valor calculado de celdas con fórmula (p.ej. =SUM).
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
    except Exception as exc:
        raise ValueError(
            "No se pudo leer el archivo. Asegurate de que sea un Excel válido "
            "(.xlsx o .xlsm)."
        ) from exc
    wb_data = None
    try:
        try:
            wb_data = openpyxl.load_workbook(filepath, data_only=True)
        except Exception:
            wb_data = None
        sheet_name = _detect_planilla_sheet(wb)
        if not sheet_name:
            return
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name] if wb_data and sheet_name in wb_data.sheetnames else None
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return

        data_start = _find_data_start(all_rows)
        header_rows = all_rows[:data_start]
        year_blocks = _detect_year_blocks(header_rows)

        all_vac: list[dict] = []
        for anio, col_offset in year_blocks:
            all_vac.extend(_parse_year_block_ws(ws, data_start, col_offset, anio, ws_data))
    finally:
        wb.close()
        if wb_data is not None:
            wb_data.close()

    if not all_vac:
        return

    anios_en_excel = {r["anio"] for r in all_vac}
    existing_vac = db.execute(
        select(Vacacion.legajo, Vacacion.anio).where(Vacacion.anio.in_(anios_en_excel))
    ).fetchall()
    existing_vac_keys = {(row.legajo, row.anio) for row in existing_vac}
    keys_vac_excel = {(r["legajo"], r["anio"]) for r in all_vac}
    result["vacaciones_nuevas"] = len(keys_vac_excel - existing_vac_keys)
    result["vacaciones_actualizadas"] = len(keys_vac_excel & existing_vac_keys)

    # Reemplazo por año: el Excel actualizado es la fuente de verdad de esos períodos.
    db.execute(delete(Vacacion).where(Vacacion.anio.in_(anios_en_excel)))
    db.bulk_insert_mappings(Vacacion, all_vac)
    db.commit()
    result["vacaciones"] = len(all_vac)


# ── Punto de entrada ────────────────────────────────────────────────────────


def _empty_import_result() -> dict:
    return {
        "registros": 0,
        "registros_nuevos": 0,
        "registros_actualizados": 0,
        "vacaciones": 0,
        "vacaciones_nuevas": 0,
        "vacaciones_actualizadas": 0,
        "fecha_min": None,
        "fecha_max": None,
        "errores": [],
    }


def import_excel(filepath: str, db: Session) -> dict:
    result = _empty_import_result()
    _import_total(filepath, db, result)
    _import_planilla(filepath, db, result)
    return result


def import_total_excel(filepath: str, db: Session) -> dict:
    """Importa horas diarias: fechas nuevas se agregan; (fecha, empleado) repetidos se pisan."""
    if not _workbook_has_hours_sheet(filepath):
        raise ValueError(
            "No se encontró una planilla de horas diarias en el archivo. "
            "Tiene que haber columnas de fecha, empleado y horas "
            "(por ejemplo total horas, hs 50%, ausente, etc.)."
        )
    result = _empty_import_result()
    _import_total(filepath, db, result)
    if result["registros"] == 0:
        raise ValueError(
            "Se detectó la planilla de horas pero no había filas válidas "
            "(fecha + empleado)."
        )
    return result
